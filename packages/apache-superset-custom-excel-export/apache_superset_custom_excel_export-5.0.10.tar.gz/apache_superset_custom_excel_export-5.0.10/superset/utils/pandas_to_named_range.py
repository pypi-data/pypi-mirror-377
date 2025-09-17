import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def write_df_to_named_range(df, wb: Workbook, named_range, include_header=True) -> Workbook:
    """
    Write a pandas DataFrame to a named range in an Excel file.
    
    Parameters:
    df (pd.DataFrame): The DataFrame to write
    excel_file (str): Path to the Excel file
    named_range (str): Name of the range to write to
    include_header (bool): Whether to include column headers
    
    Returns:
    openpyxl.Workbook: The modified workbook object
    """
    # Get the named range
    if named_range not in wb.defined_names:
        raise ValueError(f"Named range '{named_range}' not found in workbook")
    
    # Get the range reference
    range_ref = wb.defined_names[named_range]
    worksheet_name = range_ref.attr_text.split('!')[0]
    range_address = range_ref.attr_text.split('!')[1]
    
    # Get the worksheet
    ws = wb[worksheet_name]
    
    # Parse the range to get starting cell
    start_cell = range_address.split(':')[0]
    
    # Convert DataFrame to rows
    rows = dataframe_to_rows(df, index=False, header=include_header)
    
    # Write data starting from the named range's top-left cell
    start_row = ws[start_cell].row
    start_col = ws[start_cell].column
    
    for i, row in enumerate(rows):
        for j, value in enumerate(row):
            ws.cell(row=start_row + i, column=start_col + j, value=value)
    
    return wb
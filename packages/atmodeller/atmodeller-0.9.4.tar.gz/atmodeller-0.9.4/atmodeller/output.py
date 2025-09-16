#
# Copyright 2024 Dan J. Bower
#
# This file is part of Atmodeller.
#
# Atmodeller is free software: you can redistribute it and/or modify it under the terms of the GNU
# General Public License as published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# Atmodeller is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without
# even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License along with Atmodeller. If not,
# see <https://www.gnu.org/licenses/>.
#
"""Output"""

import logging
import pickle
from pathlib import Path
from typing import Any, Optional

import jax.numpy as jnp
import numpy as np
import numpy.typing as npt
import pandas as pd
from jaxmod.constants import AVOGADRO, GAS_CONSTANT
from jaxmod.units import unit_conversion
from jaxtyping import Array, ArrayLike, Bool, Float, Integer
from molmass import Formula
from openpyxl.styles import PatternFill
from scipy.constants import mega

from atmodeller import override
from atmodeller.containers import Parameters, Planet, SpeciesCollection
from atmodeller.engine_vmap import VmappedFunctions
from atmodeller.interfaces import RedoxBufferProtocol
from atmodeller.thermodata import IronWustiteBuffer
from atmodeller.type_aliases import NpArray, NpBool, NpFloat, NpInt

logger: logging.Logger = logging.getLogger(__name__)


class Output:
    """Output

    Args:
        parameters: Parameters
        solution: Solution
    """

    def __init__(self, parameters: Parameters, solution: Float[Array, " batch solution"]):
        logger.debug("Creating Output")
        self.parameters: Parameters = parameters
        self.solution: NpFloat = np.asarray(solution)
        self.vmapf: VmappedFunctions = VmappedFunctions(parameters)

        log_number_density, log_stability = np.split(self.solution, 2, axis=1)
        self.log_number_density: NpFloat = log_number_density
        # Mask stabilities that are not solved
        self.log_stability: NpFloat = np.where(
            parameters.species.active_stability, log_stability, np.nan
        )
        # Caching output to avoid recomputation
        self._cached_dict: Optional[dict[str, dict[str, NpArray]]] = None
        self._cached_dataframes: Optional[dict[str, pd.DataFrame]] = None

    @property
    def condensed_species_mask(self) -> NpBool:
        """Mask of condensed species"""
        return np.invert(self.parameters.species.gas_species_mask)

    @property
    def gas_species_mask(self) -> NpBool:
        """Mask of gas species"""
        return self.parameters.species.gas_species_mask

    @property
    def molar_mass(self) -> NpFloat:
        """Molar mass of all species"""
        return self.parameters.species.molar_masses

    @property
    def number_solutions(self) -> int:
        """Number of solutions"""
        return self.parameters.batch_size

    @property
    def planet(self) -> Planet:
        """Planet"""
        return self.parameters.planet

    @property
    def species(self) -> SpeciesCollection:
        """Species"""
        return self.parameters.species

    @property
    def temperature(self) -> NpFloat:
        """Temperature"""
        return np.asarray(self.planet.temperature)

    def activity(self) -> NpFloat:
        """Gets the activity of all species.

        Returns:
            Activity of all species
        """
        return np.exp(self.log_activity())

    def activity_without_stability(self) -> NpFloat:
        """Gets the activity without stability of all species

        Returns:
            Activity without stability of all species
        """
        return np.exp(self.log_activity_without_stability())

    def asdict(self) -> dict[str, dict[str, NpArray]]:
        """Gets all output in a dictionary, with caching.

        Returns:
            Dictionary of all output
        """
        if self._cached_dict is not None:
            logger.info("Returning cached asdict output")
            return self._cached_dict  # Return cached result

        logger.info("Computing asdict output")

        out: dict[str, dict[str, NpArray]] = {}

        # These are required for condensed and gas species
        molar_mass: NpFloat = self.species_molar_mass_expanded()
        number_density: NpFloat = self.number_density()
        activity: NpFloat = self.activity()

        gas_species_asdict = self.gas_species_asdict(molar_mass, number_density, activity)
        out |= gas_species_asdict
        out |= self.condensed_species_asdict(molar_mass, number_density, activity)
        out |= self.elements_asdict()

        out["planet"] = broadcast_arrays_in_dict(self.planet.asdict(), self.number_solutions)
        out["atmosphere"] = self.atmosphere_asdict()
        # Temperature and pressure have already been expanded to the number of solutions
        temperature: NpFloat = out["planet"]["surface_temperature"]
        pressure: NpFloat = out["atmosphere"]["pressure"]
        # Convenient to also attach temperature to the atmosphere output
        out["atmosphere"]["temperature"] = temperature
        out["raw"] = self.raw_solution_asdict()

        if "O2_g" in out:
            logger.debug("Found O2_g so back-computing log10 shift for fO2")
            log10_fugacity: NpFloat = np.log10(out["O2_g"]["fugacity"])
            buffer: RedoxBufferProtocol = IronWustiteBuffer()
            # Shift at 1 bar
            buffer_at_one_bar: NpFloat = np.asarray(buffer.log10_fugacity(temperature, 1.0))
            log10_shift_at_one_bar: NpFloat = log10_fugacity - buffer_at_one_bar
            # logger.debug("log10_shift_at_1bar = %s", log10_shift_at_one_bar)
            out["O2_g"]["log10dIW_1_bar"] = log10_shift_at_one_bar
            # Shift at actual pressure
            buffer_at_P: NpFloat = np.asarray(buffer.log10_fugacity(temperature, pressure))
            log10_shift_at_P: NpFloat = log10_fugacity - buffer_at_P
            # logger.debug("log10_shift_at_P = %s", log10_shift_at_P)
            out["O2_g"]["log10dIW_P"] = log10_shift_at_P

        # For debugging to confirm all outputs are numpy arrays
        # def find_non_numpy(d) -> None:
        #     for key, value in d.items():
        #         if isinstance(value, dict):
        #             find_non_numpy(value)
        #         else:
        #             if not isinstance(value, np.ndarray):
        #                 logger.warning("Non numpy array type found")
        #                 logger.warning("key = %s, value = %s", key, value)
        #                 logger.warning("type = %s", type(value))

        # find_non_numpy(out)

        self._cached_dict = out  # Cache result for faster re-accessing

        return out

    def atmosphere_asdict(self) -> dict[str, NpArray]:
        """Gets the atmosphere properties.

        Returns:
            Atmosphere properties
        """
        out: dict[str, NpArray] = {}

        log_number_density: Array = self.vmapf.get_log_number_density_from_log_pressure(
            jnp.log(self.total_pressure()), jnp.asarray(self.temperature)
        )
        # Must be 2-D to align arrays for computing number-density-related quantities
        number_density: NpArray = np.exp(log_number_density)[:, np.newaxis]
        molar_mass: NpArray = self.atmosphere_molar_mass()[:, np.newaxis]
        out: dict[str, NpArray] = self._get_number_density_output(
            number_density, molar_mass, "species_"
        )
        # Species mass is simply mass so rename for clarity
        out["mass"] = out.pop("species_mass")

        out["molar_mass"] = molar_mass
        # Ensure all arrays are 1-D, which is required for creating dataframes
        out = {key: value.ravel() for key, value in out.items()}

        out["pressure"] = self.total_pressure()
        out["volume"] = self.atmosphere_volume()
        out["element_number_density"] = np.sum(self.element_density_gas(), axis=1)
        out["element_number"] = out["element_number_density"] * out["volume"]
        out["element_moles"] = out["element_number"] / AVOGADRO

        return out

    def atmosphere_log_molar_mass(self) -> NpFloat:
        """Gets log molar mass of the atmosphere.

        Returns:
            Log molar mass of the atmosphere
        """
        atmosphere_log_molar_mass: Array = self.vmapf.get_atmosphere_log_molar_mass(
            jnp.asarray(self.log_number_density)
        )

        return np.asarray(atmosphere_log_molar_mass)

    def atmosphere_molar_mass(self) -> NpArray:
        """Gets the molar mass of the atmosphere.

        Returns:
            Molar mass of the atmosphere
        """
        return np.exp(self.atmosphere_log_molar_mass())

    def atmosphere_log_volume(self) -> NpFloat:
        """Gets the log volume of the atmosphere.

        Returns:
            Log volume of the atmosphere
        """
        atmosphere_log_volume: Array = self.vmapf.get_atmosphere_log_volume(
            jnp.asarray(self.log_number_density)
        )

        return np.asarray(atmosphere_log_volume)

    def atmosphere_volume(self) -> NpFloat:
        """Gets the volume of the atmosphere.

        Returns:
            Volume of the atmosphere
        """
        return np.exp(self.atmosphere_log_volume())

    def total_pressure(self) -> NpFloat:
        """Gets total pressure.

        Returns:
            Total pressure
        """
        total_pressure: Array = self.vmapf.get_total_pressure(jnp.asarray(self.log_number_density))

        return np.asarray(total_pressure)

    def condensed_species_asdict(
        self, molar_mass: NpArray, number_density: NpArray, activity: NpArray
    ) -> dict[str, dict[str, NpArray]]:
        """Gets the condensed species output as a dictionary.

        Args:
            molar_mass: Molar mass of all species
            number_density: Number density of all species
            activity: Activity of all species

        Returns:
            Condensed species output as a dictionary
        """
        molar_mass = molar_mass[:, self.condensed_species_mask]
        number_density = number_density[:, self.condensed_species_mask]
        activity = activity[:, self.condensed_species_mask]

        condensed_species: tuple[str, ...] = self.species.condensed_species_names

        out: dict[str, NpArray] = self._get_number_density_output(
            number_density, molar_mass, "total_"
        )
        out["molar_mass"] = molar_mass
        out["activity"] = activity

        split_dict: list[dict[str, NpArray]] = split_dict_by_columns(out)
        species_out: dict[str, dict[str, NpArray]] = {
            species_name: split_dict[ii] for ii, species_name in enumerate(condensed_species)
        }

        return species_out

    def elements_asdict(self) -> dict[str, dict[str, NpArray]]:
        """Gets the element properties as a dictionary.

        Returns:
            Element outputs as a dictionary
        """
        molar_mass: NpArray = self.element_molar_mass_expanded()
        atmosphere: NpArray = self.element_density_gas()
        condensed: NpArray = self.element_density_condensed()
        dissolved: NpArray = self.element_density_dissolved()
        total: NpArray = atmosphere + condensed + dissolved

        out: dict[str, NpArray] = self._get_number_density_output(
            atmosphere, molar_mass, "atmosphere_"
        )
        out |= self._get_number_density_output(condensed, molar_mass, "condensed_")
        out |= self._get_number_density_output(dissolved, molar_mass, "dissolved_")
        out |= self._get_number_density_output(total, molar_mass, "total_")

        out["molar_mass"] = molar_mass
        out["degree_of_condensation"] = out["condensed_number"] / out["total_number"]
        out["volume_mixing_ratio"] = out["atmosphere_number"] / np.sum(
            out["atmosphere_number"], axis=1, keepdims=True
        )
        out["atmosphere_ppm"] = out["volume_mixing_ratio"] * mega
        out["atmosphere_ppmw"] = (
            out["atmosphere_mass"] / np.sum(out["atmosphere_mass"], axis=1, keepdims=True) * mega
        )

        unique_elements: tuple[str, ...] = self.species.unique_elements
        if "H" in unique_elements:
            index: int = unique_elements.index("H")
            H_total_moles: NpArray = out["total_moles"][:, index]
            out["logarithmic_abundance"] = (
                np.log10(out["total_moles"] / H_total_moles[:, np.newaxis]) + 12
            )

        # logger.debug("out = %s", out)

        split_dict: list[dict[str, NpArray]] = split_dict_by_columns(out)
        # logger.debug("split_dict = %s", split_dict)

        elements_out: dict[str, dict[str, NpArray]] = {
            f"element_{element}": split_dict[ii] for ii, element in enumerate(unique_elements)
        }
        # logger.debug("elements_out = %s", elements_out)

        return elements_out

    def element_density_condensed(self) -> NpFloat:
        """Gets the number density of elements in the condensed phase.

        Returns:
            Number density of elements in the condensed phase
        """
        condensed_species_mask: NpFloat = np.where(self.condensed_species_mask, 1.0, np.nan)
        element_density: Array = self.vmapf.get_element_density(
            jnp.asarray(self.log_number_density) * condensed_species_mask
        )

        return np.asarray(element_density)

    def element_density_dissolved(self) -> NpFloat:
        """Gets the number density of elements dissolved in melt due to species solubility.

        Returns:
            Number density of elements dissolved in melt due to species solubility
        """
        element_density_dissolved: Array = self.vmapf.get_element_density_in_melt(
            jnp.asarray(self.log_number_density)
        )

        return np.asarray(element_density_dissolved)

    def element_density_gas(self) -> NpFloat:
        """Gets the number density of elements in the gas phase.

        Returns:
            Number density of elements in the gas phase
        """
        gas_species_mask: NpFloat = np.where(self.gas_species_mask, 1.0, np.nan)
        element_density: Array = self.vmapf.get_element_density(
            jnp.asarray(self.log_number_density) * gas_species_mask,
        )

        return np.asarray(element_density)

    def element_molar_mass_expanded(self) -> NpFloat:
        """Gets molar mass of elements.

        Returns:
            Molar mass of elements
        """
        unique_elements: tuple[str, ...] = self.species.unique_elements
        molar_mass: npt.ArrayLike = np.array(
            [Formula(element).mass for element in unique_elements]
        )
        molar_mass = unit_conversion.g_to_kg * molar_mass

        return np.tile(molar_mass, (self.number_solutions, 1))

    def _get_number_density_output(
        self, number_density: NpArray, molar_mass_expanded: NpArray, prefix: str = ""
    ) -> dict[str, NpArray]:
        """Gets the outputs associated with a number density.

        Args:
            number_density: Number density
            molar_mass_expanded: Molar mass associated with the number density
            prefix: Key prefix for the output. Defaults to an empty string.

        Returns
            Dictionary of output quantities
        """
        atmosphere_volume: NpArray = self.atmosphere_volume()
        # Volume must be a column vector because it multiples all elements in the row
        number: NpArray = number_density * atmosphere_volume[:, np.newaxis]
        moles: NpArray = number / AVOGADRO
        mass: NpArray = moles * molar_mass_expanded

        out: dict[str, NpArray] = {}
        out[f"{prefix}number_density"] = number_density
        out[f"{prefix}number"] = number
        out[f"{prefix}moles"] = moles
        out[f"{prefix}mass"] = mass

        return out

    def gas_species_asdict(
        self,
        molar_mass: NpArray,
        number_density: NpArray,
        activity: NpArray,
    ) -> dict[str, dict[str, NpArray]]:
        """Gets the gas species output as a dictionary.

        Args:
            molar_mass: Molar mass of all species
            number_density: Number density of all species
            activity: Activity of all species

        Returns:
            Gas species output as a dictionary
        """
        # Below are all filtered to only include the data (columns) of gas species
        molar_mass = molar_mass[:, self.gas_species_mask]
        number_density = number_density[:, self.gas_species_mask]
        activity = activity[:, self.gas_species_mask]
        dissolved_number_density: NpArray = self.species_density_in_melt()[
            :, self.gas_species_mask
        ]
        total_number_density: NpArray = number_density + dissolved_number_density
        pressure: NpArray = self.pressure()[:, self.gas_species_mask]

        gas_species: tuple[str, ...] = self.species.gas_species_names

        out: dict[str, NpArray] = {}
        out |= self._get_number_density_output(number_density, molar_mass, "atmosphere_")
        out |= self._get_number_density_output(dissolved_number_density, molar_mass, "dissolved_")
        out |= self._get_number_density_output(total_number_density, molar_mass, "total_")
        out["molar_mass"] = molar_mass
        out["volume_mixing_ratio"] = out["atmosphere_number"] / np.sum(
            out["atmosphere_number"], axis=1, keepdims=True
        )
        out["atmosphere_ppm"] = out["volume_mixing_ratio"] * mega
        out["atmosphere_ppmw"] = (
            out["atmosphere_mass"] / np.sum(out["atmosphere_mass"], axis=1, keepdims=True) * mega
        )
        out["pressure"] = pressure
        out["fugacity"] = activity
        out["fugacity_coefficient"] = activity / pressure
        out["dissolved_ppmw"] = self.species_ppmw_in_melt()

        split_dict: list[dict[str, NpArray]] = split_dict_by_columns(out)
        species_out: dict[str, dict[str, NpArray]] = {
            species_name: split_dict[ii] for ii, species_name in enumerate(gas_species)
        }

        return species_out

    def log_activity(self) -> NpFloat:
        """Gets log activity of all species.

        This is usually what the user wants when referring to activity because it includes a
        consideration of species stability

        Returns:
            Log activity of all species
        """
        log_activity_without_stability: NpFloat = self.log_activity_without_stability()
        log_activity_with_stability: NpFloat = log_activity_without_stability - np.exp(
            self.log_stability
        )
        # Now select the appropriate activity for each species, depending if stability is relevant.
        condition_broadcasted = np.broadcast_to(
            self.parameters.species.active_stability, log_activity_without_stability.shape
        )
        # logger.debug("condition_broadcasted = %s", condition_broadcasted)

        log_activity: NpFloat = np.where(
            condition_broadcasted,
            log_activity_with_stability,
            log_activity_without_stability,
        )

        return log_activity

    def log_activity_without_stability(self) -> NpFloat:
        """Gets log activity without stability of all species.

        Returns:
            Log activity without stability
        """
        log_activity: Array = self.vmapf.get_log_activity(jnp.asarray(self.log_number_density))

        return np.asarray(log_activity)

    def number_density(self) -> NpFloat:
        r"""Gets number density of all species.

        Returns:
            Number density in :math:`\mathrm{molecules}\, \mathrm{m}^{-3}`
        """
        return np.exp(self.log_number_density)

    def reaction_mask(self) -> NpBool:
        """Gets the reaction mask of the residual array.

        Returns:
            Reaction mask of the residual array
        """
        reaction_mask: Bool[Array, "..."] = self.vmapf.get_reactions_only_mask()

        return np.asarray(reaction_mask, dtype=bool)

    def species_molar_mass_expanded(self) -> NpFloat:
        """Gets molar mass of all species in an expanded array.

        Returns:
            Molar mass of all species in an expanded array.
        """
        return np.tile(self.molar_mass, (self.number_solutions, 1))

    def pressure(self) -> NpFloat:
        """Gets pressure of species in bar.

        This will compute pressure of all species, including condensates, for simplicity.

        Returns:
            Pressure of species in bar
        """
        pressure: Array = self.vmapf.get_pressure_from_log_number_density(
            jnp.asarray(self.log_number_density)
        )

        return np.asarray(pressure)

    def quick_look(self) -> dict[str, ArrayLike]:
        """Quick look at the solution

        Provides a quick first glance at the output with convenient units and to ease comparison
        with test or benchmark data.

        Returns:
            Dictionary of the solution
        """
        out: dict[str, ArrayLike] = {}

        for nn, species_ in enumerate(self.species):
            pressure: NpArray = self.pressure()[:, nn]
            activity: NpArray = self.activity()[:, nn]
            out[species_.name] = pressure
            out[f"{species_.name}_activity"] = activity

        return {key: np.squeeze(value) for key, value in out.items()}

    def raw_solution_asdict(self) -> dict[str, NpArray]:
        """Gets the raw solution.

        Returns:
            Dictionary of the raw solution
        """
        raw_solution: dict[str, NpArray] = {}

        species_names: tuple[str, ...] = self.species.species_names

        for ii, species_name in enumerate(species_names):
            raw_solution[species_name] = self.log_number_density[:, ii]
            raw_solution[f"{species_name}_stability"] = self.log_stability[:, ii]

        # Remove keys where the array values are all nan
        for key in list(raw_solution.keys()):
            if np.all(np.isnan(raw_solution[key])):
                raw_solution.pop(key)

        return raw_solution

    def residual_asdict(self) -> dict[int, NpFloat]:
        """Gets the residual.

        Returns:
            Dictionary of the residual
        """
        residual: Array = self.vmapf.objective_function(jnp.asarray(self.solution))

        out: dict[int, NpArray] = {}
        for ii in range(residual.shape[1]):
            out[ii] = np.asarray(residual[:, ii])

        return out

    def species_density_in_melt(self) -> NpFloat:
        """Gets species number density in the melt.

        Returns:
            Species number density in the melt
        """
        species_density_in_melt: Array = self.vmapf.get_species_density_in_melt(
            jnp.asarray(self.log_number_density)
        )

        return np.asarray(species_density_in_melt)

    def species_ppmw_in_melt(self) -> NpFloat:
        """Gets species ppmw in the melt.

        Return:
            Species ppmw in the melt
        """
        species_ppmw_in_melt: Array = self.vmapf.get_species_ppmw_in_melt(
            jnp.asarray(self.log_number_density)
        )

        return np.asarray(species_ppmw_in_melt)

    def stability(self) -> NpFloat:
        """Gets stability of relevant species.

        Returns:
            Stability of relevant species
        """
        return np.exp(self.log_stability)

    def to_dataframes(self) -> dict[str, pd.DataFrame]:
        """Gets the output in a dictionary of dataframes.

        Returns:
            Output in a dictionary of dataframes
        """
        if self._cached_dataframes is not None:
            logger.debug("Returning cached to_dataframes output")
            dataframes: dict[str, pd.DataFrame] = self._cached_dataframes  # Return cached result
        else:
            logger.info("Computing to_dataframes output")
            dataframes = nested_dict_to_dataframes(self.asdict())
            self._cached_dataframes = dataframes
            # logger.debug("to_dataframes = %s", self._cached_dataframes)

        return dataframes

    def to_excel(self, file_prefix: Path | str = "new_atmodeller_out") -> None:
        """Writes the output to an Excel file.

        Args:
            file_prefix: Prefix of the output file. Defaults to new_atmodeller_out.
        """
        logger.info("Writing output to excel")
        out: dict[str, pd.DataFrame] = self.to_dataframes()
        output_file: Path = Path(f"{file_prefix}.xlsx")

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for df_name, df in out.items():
                df.to_excel(writer, sheet_name=df_name, index=True)

        logger.info("Output written to %s", output_file)

    def to_pickle(self, file_prefix: Path | str = "new_atmodeller_out") -> None:
        """Writes the output to a pickle file.

        Args:
            file_prefix: Prefix of the output file. Defaults to new_atmodeller_out.
        """
        logger.info("Writing output to pickle")
        out: dict[str, pd.DataFrame] = self.to_dataframes()
        output_file: Path = Path(f"{file_prefix}.pkl")

        with open(output_file, "wb") as handle:
            pickle.dump(out, handle, protocol=pickle.HIGHEST_PROTOCOL)

        logger.info("Output written to %s", output_file)


class OutputDisequilibrium(Output):
    """Output disequilibrium calculations

    Args:
        parameters: Parameters
        solution: Solution
    """

    @override
    def asdict(self) -> dict[str, dict[str, NpArray]]:
        """All outputs in a dictionary, with caching.

        Additionally includes the disequilibrium group, compared to the base class.

        Returns:
            Dictionary of all output
        """
        out: dict[str, dict[str, NpArray]] = super().asdict()

        out["disequilibrium"] = self.disequilibrium_asdict()

        self._cached_dict = out  # Re-cache result for faster re-accessing

        return out

    def disequilibrium_asdict(self) -> dict[str, NpArray]:
        """Gets the reaction disequilibrium as a dictionary.

        Returns:
            Reaction disequilibrium as a dictionary
        """
        reaction_mask: NpBool = self.reaction_mask()
        residual: NpFloat = np.asarray(self.vmapf.objective_function(jnp.asarray(self.solution)))

        # Number of True entries per row (must be same for all rows)
        n_cols: NpInt = reaction_mask.sum(axis=1)[0]
        # logger.debug("n_cols = %s", n_cols)
        # Convert boolean mask to sorted column indices for each row
        col_indices: NpInt = np.argsort(~reaction_mask, axis=1)[:, :n_cols]
        # logger.debug("col_indices = %s", col_indices)
        # Gather the True entries in order
        compressed: NpFloat = np.take_along_axis(residual, col_indices, axis=1)
        # logger.debug("compressed = %s", compressed)

        # To compute the limiting reactant/product in each reaction we need to know the
        # availability of each species. We will ignore condensates later because their stability
        # criteria prevents a simple calculation of what is limiting the reaction.
        number_density: NpFloat = self.number_density()
        # logger.debug("number_density = %s", number_density)
        number_fraction: NpFloat = number_density / np.sum(number_density, axis=1, keepdims=True)
        # logger.debug("number_fraction = %s", number_fraction)
        reaction_matrix: NpFloat = self.parameters.species.reaction_matrix
        # logger.debug("reaction_matrix = %s", reaction_matrix)

        out: dict[str, NpArray] = {}

        for jj in range(n_cols):
            logger.debug("Working on reaction %d", jj)
            per_mole_of_reaction: NpFloat = compressed[:, jj] * GAS_CONSTANT * self.temperature
            stoich: NpFloat = reaction_matrix[jj]
            # logger.debug("stoich = %s", stoich)

            # Normalised ratios for limiting species (ignore divide-by-zero warnings)
            with np.errstate(divide="ignore"):
                ratios: NpFloat = np.where(stoich != 0, number_fraction / stoich, np.nan)
            limiting: NpFloat = np.full_like(per_mole_of_reaction, np.nan)

            # Backward-favoured: products limit
            mask_back: NpBool = per_mole_of_reaction > 0
            if np.any(mask_back):
                limiting[mask_back] = np.nanmin(ratios[mask_back][:, stoich > 0], axis=1)

            # Forward-favoured: reactants limit
            mask_fwd: NpBool = ~mask_back
            if np.any(mask_fwd):
                # Limiting species is the largest negative ratio among reactants (closest to zero)
                limiting[mask_fwd] = np.nanmax(ratios[mask_fwd][:, stoich < 0], axis=1)

            # Compute the energy per mole of atmosphere
            energy_per_mol_atmosphere: NpFloat = per_mole_of_reaction * limiting

            out[f"Reaction_{jj}"] = per_mole_of_reaction
            if self.species.gas_only:
                out[f"Reaction_{jj}_per_atmosphere"] = energy_per_mol_atmosphere

        return out


class OutputSolution(Output):
    """Output equilibrium solution(s)

    Args:
        parameters: Parameters
        solution: Solution
        solver_status: Solver status
        solver_steps: Number of solver steps
        solver_attempts: Number of solver attempts (multistart)
    """

    def __init__(
        self,
        parameters: Parameters,
        solution: Float[Array, "batch solution"],
        solver_status: Bool[Array, "..."],
        solver_steps: Integer[Array, "..."],
        solver_attempts: Integer[Array, "..."],
    ):
        super().__init__(parameters, solution)
        self._solver_status: NpBool = np.asarray(solver_status)
        self._solver_steps: NpInt = np.asarray(solver_steps)
        self._solver_attempts: NpInt = np.asarray(solver_attempts)

    @override
    def asdict(self) -> dict[str, dict[str, NpArray]]:
        """All outputs in a dictionary, with caching.

        Returns:
            Dictionary of all output
        """
        out: dict[str, dict[str, NpArray]] = super().asdict()

        # Temperature and pressure have already been expanded to the number of solutions
        temperature: NpFloat = out["planet"]["surface_temperature"]
        pressure: NpFloat = out["atmosphere"]["pressure"]

        out["constraints"] = {}
        out["constraints"] |= broadcast_arrays_in_dict(
            self.parameters.mass_constraints.asdict(), self.number_solutions
        )
        out["constraints"] |= broadcast_arrays_in_dict(
            self.parameters.fugacity_constraints.asdict(temperature, pressure),
            self.number_solutions,
        )
        out["constraints"] |= broadcast_arrays_in_dict(
            self.parameters.total_pressure_constraint.asdict(), self.number_solutions
        )

        out["residual"] = self.residual_asdict()  # type: ignore since keys are int

        out["solver"] = {
            "status": self._solver_status,
            "steps": self._solver_steps,
            "attempts": self._solver_attempts,
        }

        self._cached_dict = out  # Re-cache result for faster re-accessing

        return out

    @override
    def to_dataframes(self, drop_unsuccessful: bool = False) -> dict[str, pd.DataFrame]:
        """Gets the output in a dictionary of dataframes.

        Args:
            drop_unsuccessful: Drop models that did not solve. Defaults to False.

        Returns:
            Output in a dictionary of dataframes
        """
        dataframes: dict[str, pd.DataFrame] = super().to_dataframes()

        if drop_unsuccessful:
            logger.info("Dropping models that did not solve")
            dataframes: dict[str, pd.DataFrame] = self._drop_unsuccessful_solves(dataframes)

        return dataframes

    @override
    def to_excel(
        self, file_prefix: Path | str = "new_atmodeller_out", drop_unsuccessful: bool = False
    ) -> None:
        """Writes the output to an Excel file.

        Compared to the base class, this highlights rows where the solver failed to find a
        a solution if `drop_successful = False`.

        Args:
            file_prefix: Prefix of the output file. Defaults to new_atmodeller_out.
            drop_unsuccessful: Drop models that did not solve. Defaults to False.
        """
        logger.info("Writing output to excel")
        out: dict[str, pd.DataFrame] = self.to_dataframes(drop_unsuccessful)
        output_file: Path = Path(f"{file_prefix}.xlsx")

        # Convenient to highlight rows where the solver failed to find a solution for follow-up
        # analysis. Define a fill colour for highlighting rows (e.g., yellow)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Get the indices where the successful_solves mask is False
        unsuccessful_indices: NpArray = np.where(
            np.array(self._solver_status) == False  # noqa: E712
        )[0]

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for df_name, df in out.items():
                df.to_excel(writer, sheet_name=df_name, index=True)
                sheet = writer.sheets[df_name]

                # Apply highlighting to the rows where the solver failed to find a solution
                for idx in unsuccessful_indices:
                    # Highlight the entire row (starting from index 2 to skip header row)
                    for col in range(1, len(df.columns) + 2):
                        # row=idx+2 because Excel is 1-indexed and row 1 is the header
                        cell = sheet.cell(row=idx + 2, column=col)
                        cell.fill = highlight_fill

        logger.info("Output written to %s", output_file)

    @override
    def to_pickle(
        self, file_prefix: Path | str = "new_atmodeller_out", drop_unsuccessful: bool = False
    ) -> None:
        """Writes the output to a pickle file.

        Args:
            file_prefix: Prefix of the output file. Defaults to new_atmodeller_out.
            drop_unsuccessful: Drop models that did not solve. Defaults to False.
        """
        logger.info("Writing output to pickle")
        out: dict[str, pd.DataFrame] = self.to_dataframes(drop_unsuccessful)
        output_file: Path = Path(f"{file_prefix}.pkl")

        with open(output_file, "wb") as handle:
            pickle.dump(out, handle, protocol=pickle.HIGHEST_PROTOCOL)

        logger.info("Output written to %s", output_file)

    def _drop_unsuccessful_solves(
        self, dataframes: dict[str, pd.DataFrame]
    ) -> dict[str, pd.DataFrame]:
        """Drops unsuccessful solves.

        Args:
            dataframes: Dataframes from which to drop unsuccessful models

        Returns:
            Dictionary of dataframes without unsuccessful models
        """
        return {key: df.loc[self._solver_status] for key, df in dataframes.items()}


def broadcast_arrays_in_dict(some_dict: dict[str, NpArray], shape: int) -> dict[str, NpArray]:
    """Gets a dictionary of broadcasted arrays.

    Args:
        some_dict: Some dictionary
        size: Shape (size) of the desired array

    Returns:
        A dictionary with broadcasted arrays
    """
    expanded_dict: dict[str, NpArray] = {}
    for key, value in some_dict.items():
        expanded_dict[key] = np.broadcast_to(value, shape)

    return expanded_dict


def split_dict_by_columns(dict_to_split: dict[str, NpArray]) -> list[dict[str, NpArray]]:
    """Splits a dictionary based on columns in the values.

    Args:
        dict_to_split: A dictionary to split

    Returns:
        A list of dictionaries split by column
    """
    # Assume all arrays have the same number of columns
    first_key: str = next(iter(dict_to_split))
    num_columns: int = dict_to_split[first_key].shape[1]

    # Preallocate list of dicts
    split_dicts: list[dict] = [{} for _ in range(num_columns)]

    for key, array in dict_to_split.items():
        for i in range(num_columns):
            split_dicts[i][key] = array[:, i]

    return split_dicts


def nested_dict_to_dataframes(nested_dict: dict[str, dict[str, Any]]) -> dict[str, pd.DataFrame]:
    """Creates a dictionary of dataframes from a nested dictionary.

    Args:
        nested_dict: A nested dictionary

    Returns:
        A dictionary of dataframes
    """
    dataframes: dict[str, pd.DataFrame] = {}

    for outer_key, inner_dict in nested_dict.items():
        # Convert inner dictionary to DataFrame
        df: pd.DataFrame = pd.DataFrame(inner_dict)
        dataframes[outer_key] = df

    return dataframes

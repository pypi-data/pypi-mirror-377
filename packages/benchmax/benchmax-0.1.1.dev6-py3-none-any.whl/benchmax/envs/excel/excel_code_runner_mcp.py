from pathlib import Path
import platform
import shutil
from fastmcp import FastMCP
import tempfile
import subprocess
import os
import sys
from openpyxl import load_workbook

mcp = FastMCP(
    name="ExcelCodeRunner",
    instructions="This server provides a tool for running Python code to manipulate Excel files."
)

WHITE_LIKE_COLORS = [
    "00000000",
    "FFFFFFFF",
    "FFFFFF00",
]

def evaluate_excel(excel_path: str):
    """
    Evaluate Python code that manipulates an Excel file using xlwings.
    """
    import xlwings
    # Use LibreOffice for Linux
    if platform.system() == "Linux":
        evaluate_excel_libre(excel_path)
        return
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(excel_path)
    excel_book.save()
    excel_book.close()
    excel_app.quit()

def evaluate_excel_libre(excel_path: str) -> None:
    """
    Force‑recalculate in place under Linux using LibreOffice.
    Raises subprocess.CalledProcessError if soffice exits abnormally.
    """
    tmp_outdir = tempfile.mkdtemp(prefix="lo_convert_")
    cmd = [
        "soffice",
        "--headless",
        "--nologo", "--nofirststartwizard", "--norestore",
        "--calc",
        "--convert-to", "xlsx",
        "--outdir", tmp_outdir,
        os.path.abspath(excel_path),
    ]
    lo_home = Path(tempfile.mkdtemp(prefix="lo_profile_"))
    env = dict(os.environ, HOME=str(lo_home))
    try:
        subprocess.run(
            cmd, check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            env=env,
        )
         # Determine the converted file name (same base name, .xlsx extension)
        base_name = os.path.splitext(os.path.basename(excel_path))[0] + ".xlsx"
        converted_path = os.path.join(tmp_outdir, base_name)
        # Overwrite the original file with the converted one
        shutil.move(converted_path, excel_path)
    finally:
        # Clean up the temp folder
        shutil.rmtree(tmp_outdir, ignore_errors=True)
        pass

def excel_to_str_repr(excel_path: str, evaluate_formulas = False) -> str:
    # Load workbook twice: data_only=True to get the evaluated values,
    # and data_only=False to get the formulas and styles.
    if evaluate_formulas:
        evaluate_excel(excel_path)

    wb_evaluated = load_workbook(excel_path, data_only=True)
    wb_raw = load_workbook(excel_path, data_only=False)

    result = []

    for sheet_name in wb_evaluated.sheetnames:
        sheet_evaluated = wb_evaluated[sheet_name]
        sheet_raw = wb_raw[sheet_name]

        sheet_result = f"Sheet Name: {sheet_name}"
        result.append(sheet_result)

        for row_evaluated, row_raw in zip(sheet_evaluated.iter_rows(), sheet_raw.iter_rows()):
            is_row_empty = True

            for cell_evaluated, cell_raw in zip(row_evaluated, row_raw):
                is_default_background = True
                style = []

                if (
                    cell_raw.fill.start_color.index != "00000000"
                    and type(cell_raw.fill.start_color.rgb) is str
                    and cell_raw.fill.start_color.rgb not in WHITE_LIKE_COLORS
                ):
                    is_default_background = False
                    style.append(f"bg:{cell_raw.fill.start_color.rgb}")
                if cell_raw.font.color and cell_raw.font.color.index != 1 and type(cell_raw.font.color.rgb) is str:
                    style.append(f"color:{cell_raw.font.color.rgb}")
                if cell_raw.font.bold:
                    style.append("bold")
                if cell_raw.font.italic:
                    style.append("italic")
                if cell_raw.font.underline:
                    style.append("underline")

                display_value = cell_evaluated.value
                if cell_raw.data_type == "f":
                    cell_raw_val = cell_raw.value
                    if type(cell_raw_val) is not str:
                        cell_raw_val = cell_raw.value.text # type: ignore
                    display_value = f"{cell_raw_val} -> {cell_evaluated.value}"

                coords = cell_evaluated.coordinate
                
                if display_value is None and not is_default_background:
                    # If cell is empty but has background color, still include it
                    result.append(f"{coords}: null [{', '.join(style)}]")
                    is_row_empty = False
                elif display_value:
                    style_str = f" [{", ".join(style)}]" if style else ""
                    result.append(f"{coords}: {display_value}{style_str}")
                    is_row_empty = False
            if not is_row_empty:
                result.append("") # Newline after each row

    return "\n".join(result)

def run_excel_code_impl(
    python_code: str,
    output_excel_path: str
) -> str:
    """
    Run Python code which should use openpyxl to manipulate an Excel file.
    Call load_workbook with the input excel path as specified by the user.
    Remember to save the workbook to the output path that you specified and then call close() so you do not overwrite the input file.

    If code executes with no errors, return the string representation of the Excel file with styles.
    If there are errors, return an error message.
    """
    code_path = "script.py"
    # Write the user code to a file
    with open(code_path, "w") as f:
        f.write(python_code)
    try:
        subprocess.run(
            [sys.executable, code_path],
            check=True,
            capture_output=True,
            timeout=60
        )
    except subprocess.CalledProcessError as e:
        return f"ERROR: User code failed: {e.stderr.decode()}"
    except Exception as e:
        return f"ERROR: Error running user code: {str(e)}"
    # Convert the manipulated Excel file to JSON with styles
    excel_str = excel_to_str_repr(output_excel_path)
    return excel_str

@mcp.tool
def run_excel_code(
    python_code: str,
    output_excel_path: str
) -> str:
    """
    Run Python code which should use openpyxl to manipulate an Excel file.
    If code executes with no errors, returns the string representation of the Excel file with styles.
    If there are errors, return an error message.
    """
    return run_excel_code_impl(python_code, output_excel_path)

if __name__ == "__main__":
    mcp.run(show_banner=False)
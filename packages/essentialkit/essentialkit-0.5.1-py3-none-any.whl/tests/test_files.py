import pytest
import tempfile

from openpyxl import Workbook
from assertpy import assert_that
from essentialkit.files import *
from pyparsing import ParseSyntaxException
from pyhocon import HOCONConverter, ConfigFactory


@pytest.fixture
def valid_json_file(tmp_path: Path) -> Path:
    file_path = tmp_path / "data.json"
    data = {"key": "value"}
    with open(file_path, "w") as file:
        json.dump(data, file)
    return file_path


@pytest.fixture
def invalid_file(tmp_path: Path) -> Path:
    file_path = tmp_path / "invalid_data"
    with open(file_path, "w") as file:
        file.write("invalid content")
    return file_path


@pytest.fixture
def valid_txt_file(tmp_path: Path) -> Path:
    file_path = tmp_path / "data.txt"
    with open(file_path, "w") as file:
        file.write("This is text file")
    return file_path


@pytest.fixture
def valid_hocon_file(tmp_path: Path) -> Path:
    file_path = tmp_path / "data.conf"
    data = {"key": "value"}
    data_parsed = ConfigFactory.from_dict(data)
    with open(file_path, "w") as file:
        file.write(HOCONConverter.to_hocon(data_parsed))
    return file_path


@pytest.fixture
def output_path(tmp_path: Path) -> Path:
    return tmp_path / "test_output.json"


@pytest.fixture
def create_tmp_directory(tmp_path: Path) -> Path:
    directory_path = tmp_path / "test_folder"
    directory_path.mkdir()

    sub_directory_path = directory_path / "sub_directory_path"
    sub_directory_path.mkdir()

    file_path = directory_path / "directory_file.txt"
    with open(file_path, "w") as file:
        file.write("Content in file 1")

    sub_file_path = sub_directory_path / "sub_directory_file.txt"
    with open(sub_file_path, "w") as file:
        file.write("Content in file 2")

    return directory_path


def test_receive_a_valid_directory_with_files(create_tmp_directory: Path):
    directory_path = create_tmp_directory
    result = list_files(directory_path)
    assert_that(result[0].name).is_equal_to("directory_file.txt")
    assert_that(result[1].name).is_equal_to("sub_directory_file.txt")
    for file_path in result:
        file_path.is_file()


def test_receive_an_empty_directory(tmp_path: Path):
    empty_directory = tmp_path / "empty_folder"
    empty_directory.mkdir()
    assert_that(list_files(empty_directory)).is_equal_to([])


def test_invalid_directory_path():
    invalid_path = Path("nonexistent_folder")
    assert_that(list_files).raises(OSError).when_called_with(invalid_path)


def test_file_path(valid_txt_file: Path):
    assert_that(list_files).raises(OSError).when_called_with(valid_txt_file)


def test_read_json_receive_a_valid_json_file(valid_json_file):
    expected_data = {"key": "value"}
    assert_that(read_json(valid_json_file)).is_equal_to(expected_data)


def test_read_json_receive_a_invalid_json_file(invalid_file):
    assert_that(read_json).raises(json.JSONDecodeError).when_called_with(invalid_file)


def test_read_json_receive_json_nonexistent_file():
    assert_that(read_json).raises(FileNotFoundError).when_called_with("nonexistent_file.json")


def test_read_json_receive_receive_a_txt_file(valid_txt_file):
    assert_that(read_json).raises(json.JSONDecodeError).when_called_with(valid_txt_file)


def test_write_json_data_successfully(output_path):
    test_data = {"key1": "value1", "key2": 2, "key3": [1, 2, 3]}
    write_json(test_data, output_path)
    with open(output_path, "r") as file:
        written_data = json.load(file)
    assert_that(written_data).is_equal_to(test_data)


def test_write_json_data_sorting_keys(output_path):
    test_data = {"b": 2, "a": 1}
    write_json(test_data, output_path, sort_keys=True)
    with open(output_path, "r") as file:
        written_data = json.load(file)
    assert_that(written_data).is_equal_to(test_data)


def test_write_json_invalid_input(output_path):
    assert_that(write_json).raises(TypeError).when_called_with("not a dict")


def test_write_json_invalid_output_path():
    invalid_path = "/invalid/output.json"
    assert_that(write_json).raises(FileNotFoundError).when_called_with(
        {"key": "value"}, invalid_path
    )


def test_read_hocon_receive_a_valid_hocon_file(valid_hocon_file):
    expected_data = {"key": "value"}
    assert_that(read_hocon(valid_hocon_file, True)).is_equal_to(expected_data)


def test_read_hocon_receive_a_invalid_hocon_file(invalid_file):
    assert_that(read_hocon).raises(ParseSyntaxException).when_called_with(
        invalid_file, False
    )


def test_read_hocon_receive_nonexistent_file():
    assert_that(read_hocon).raises(FileNotFoundError).when_called_with(
        "nonexistent_file.conf", False
    )


def test_read_hocon_receive_a_txt_file(valid_txt_file):
    assert_that(read_hocon).raises(ParseSyntaxException).when_called_with(
        valid_txt_file, False
    )


def test_iterate_hocon_case_str_int_bool():
    hocon_string = """
        config {
            database {
                host = "localhost"
                port = 5432
            }
            debug = true
        }
    """
    config = ConfigFactory.parse_string(hocon_string)
    result = list(iterate_hocon(config))
    expected = [
        ("config.database.host", "localhost"),
        ("config.database.port", 5432),
        ("config.debug", True)
    ]
    assert_that(result).is_equal_to(expected)


def test_iterate_hocon_case_list():
    hocon_string = """
        users = ["Alice", "Bob", "Charlie"]
    """
    config = ConfigFactory.parse_string(hocon_string)
    result = list(iterate_hocon(config))
    expected = [
        ("users[0]", "Alice"),
        ("users[1]", "Bob"),
        ("users[2]", "Charlie")
    ]
    assert_that(result).is_equal_to(expected)


def test_iterate_hocon_case_nested_lists_and_objects():
    hocon_string = """
        config {
            list = [
                { id = 1, name = "Item1" },
                { id = 2, name = "Item2" }
            ]
        }
    """
    config = ConfigFactory.parse_string(hocon_string)
    result = list(iterate_hocon(config))
    expected = [
        ("config.list[0].id", 1),
        ("config.list[0].name", "Item1"),
        ("config.list[1].id", 2),
        ("config.list[1].name", "Item2")
    ]
    assert_that(result).is_equal_to(expected)


def test_iterate_hocon_case_empty_object():
    hocon_string = """
        empty_config = {}
    """
    config = ConfigFactory.parse_string(hocon_string)
    result = list(iterate_hocon(config))
    expected = []
    assert_that(result).is_equal_to(expected)


def test_iterate_hocon_case_empty_list():
    hocon_string = """
        empty_list = []
    """
    config = ConfigFactory.parse_string(hocon_string)
    result = list(iterate_hocon(config))
    expected = []
    assert_that(result).is_equal_to(expected)


def test_iterate_hocon_case_complex_config():
    hocon_string = """
        config {
            nested {
                key1 = "value1"
                key2 {
                    subkey = "value2"
                }
            }
            list = [
                "A", "B", { nestedKey = "C" }
            ]
        }
    """
    config = ConfigFactory.parse_string(hocon_string)
    result = list(iterate_hocon(config))
    expected = [
        ("config.nested.key1", "value1"),
        ("config.nested.key2.subkey", "value2"),
        ("config.list[0]", "A"),
        ("config.list[1]", "B"),
        ("config.list[2].nestedKey", "C")
    ]
    assert_that(result).is_equal_to(expected)


def test_write_hocon_data_successfully(output_path):
    test_data = {"key1": "value1", "key2": 2, "key3": [1, 2, 3]}
    write_hocon(test_data, output_path)
    written_data = ConfigFactory.parse_file(output_path)
    assert_that(written_data).is_equal_to(test_data)


def test_write_hocon_invalid_input(output_path):
    assert_that(write_hocon).raises(TypeError).when_called_with("not a dict")


def test_write_hocon_invalid_output_path():
    invalid_path = "/invalid/output.json"
    assert_that(write_hocon).raises(FileNotFoundError).when_called_with(
        {"key": "value"}, invalid_path
    )


@pytest.fixture
def temp_excel_file():
    # Create a temp Excel file with one sheet and some dummy data
    #       A       B
    # 1    ID    Name
    # 2     1   Alice
    # 3     2     Bob
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["A2"] = 1
    ws["B2"] = "Alice"
    ws["A3"] = 2
    ws["B3"] = "Bob"

    # Save to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    yield Path(temp_file.name)

    # Cleanup
    temp_file.close()
    Path(temp_file.name).unlink()


def test_update_column_success(temp_excel_file):
    values = ["Engineer", "Manager"]
    update_excel_column(
        excel_path=temp_excel_file,
        sheet_name="Sheet1",
        column_letter="C",
        start_row=2,
        values=values
    )

    wb = load_workbook(temp_excel_file)
    ws = wb["Sheet1"]

    assert_that(ws["C1"].value).is_none()
    assert_that(ws["C2"].value).is_equal_to("Engineer")
    assert_that(ws["C3"].value).is_equal_to("Manager")


def test_update_empty_values(temp_excel_file):
    values = []
    update_excel_column(
        excel_path=temp_excel_file,
        sheet_name="Sheet1",
        column_letter="D",
        start_row=2,
        values=values
    )

    wb = load_workbook(temp_excel_file)
    ws = wb["Sheet1"]

    assert_that(ws["D2"].value).is_none()
    assert_that(ws["D3"].value).is_none()


def test_missing_sheet(temp_excel_file):
    with pytest.raises(KeyError):
        update_excel_column(
            excel_path=temp_excel_file,
            sheet_name="WrongSheet",
            column_letter="B",
            start_row=2,
            values=["X", "Y"]
        )


def test_file_not_found():
    fake_path = Path("/fake/path/nonexistent.xlsx")
    with pytest.raises(FileNotFoundError):
        update_excel_column(
            excel_path=fake_path,
            sheet_name="Sheet1",
            column_letter="A",
            start_row=1,
            values=["Test"]
        )


def test_partial_overwrite(temp_excel_file):
    values = ["Updated"]
    update_excel_column(
        excel_path=temp_excel_file,
        sheet_name="Sheet1",
        column_letter="B",
        start_row=3,
        values=values
    )

    wb = load_workbook(temp_excel_file)
    ws = wb["Sheet1"]

    assert_that(ws["B2"].value).is_equal_to("Alice")
    assert_that(ws["B3"].value).is_equal_to("Updated")

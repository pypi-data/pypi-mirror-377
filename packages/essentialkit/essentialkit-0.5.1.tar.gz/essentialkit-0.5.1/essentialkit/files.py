import json

from pathlib import Path
from openpyxl import load_workbook
from pyhocon import HOCONConverter, ConfigTree
from pyhocon.config_parser import STR_SUBSTITUTION, ConfigFactory


def list_files(path: Path) -> list[Path]:
    """
    Get a list of all file paths within a specified folder and its subdirectories.

    :param path: The path to the directory to explore
    :return: A list of file paths.
    """
    if not path.exists():
        raise OSError(f"The specified path '{path}' does not exist.")
    if not path.is_dir():
        raise OSError(f"The specified path '{path}' is not a directory.")
    return [subpath for subpath in path.rglob('*') if subpath.is_file()]


def read_json(path: Path) -> dict:
    """
    Read and parse a JSON file from the given path and return its contents as a Python dictionary.

    :param path: the file path to the JSON file to be read and parsed
    :return: a python dictionary containing the parsed JSON data
    """
    with open(path, mode="r", encoding="utf-8") as file:
        return json.load(file)


def write_json(data: dict, output_path: Path, indent: int = 4, sort_keys=False) -> None:
    """
    This function serializes a Python dictionary into JSON format and writes it to the specified
    file.

    :param data: a dictionary containing data to be written to the JSON file
    :param output_path: the file path where the JSON data will be written
    :param indent: the number of spaces used for indentation in the JSON file (default is 4)
    :param sort_keys: whether to sort the keys of the JSON file alphabetically (default is False)
    :return: None
    """
    with open(output_path, 'w') as file:
        json.dump(data, file, indent=indent, sort_keys=sort_keys)


def read_hocon(path: Path, replace_env_variables_as_str: bool) -> ConfigTree:
    """
    Read and parse a HOCON file from the given path and return its contents as a Python dictionary.

    :param path: the file path to the HOCON file to be read and parsed
    :param replace_env_variables_as_str: true cast env variables to string
    :return: a python dictionary containing the parsed HOCON data
    """
    if replace_env_variables_as_str:
        return ConfigFactory.parse_file(path, resolve=False, unresolved_value=STR_SUBSTITUTION)
    return ConfigFactory.parse_file(path, resolve=False)


def iterate_hocon(config: ConfigTree, parent_key: str = ""):
    """
    This function traverses the HOCON configuration structure and yields key-value pairs.
    Nested dictionaries and lists are flattened into dot (`.`) and index (`[i]`) notation.

    :param config: The HOCON configuration object to iterate over
    :param parent_key: The hierarchical key prefix used for recursion. Defaults to an empty string
    :return: A tuple containing the full hierarchical key as a string and its corresponding value
    """
    if isinstance(config, ConfigTree):
        for key, value in config.items():
            full_key = f"{parent_key}.{key}" if parent_key else key
            yield from iterate_hocon(value, full_key)
    elif isinstance(config, list):
        for index, item in enumerate(config):
            full_key = f"{parent_key}[{index}]"
            yield from iterate_hocon(item, full_key)
    else:
        yield parent_key, config


def write_hocon(data: dict, output_path: Path, indent: int = 2, compact=True) -> None:
    """
    This function serializes a Python dictionary into HOCON format and writes it to the specified
    file.

    :param data: a dictionary containing data to be written to the HOCON file
    :param output_path: the file path where the HOCON data will be written
    :param indent: the number of spaces used for indentation in the JSON file (default is 2)
    :param compact: whether the key of the dictionary should be compacted (default is True)
    :return: None
    """
    data_parsed = ConfigFactory.from_dict(data)
    with open(output_path, 'w') as conf_file:
        conf_file.write(HOCONConverter.to_hocon(data_parsed, compact=compact, indent=indent))


def update_excel_column(
        excel_path: Path,
        sheet_name: str,
        column_letter: str,
        start_row: int,
        values: list
) -> None:
    """
    Update a specific column in an Excel sheet with values from a list
    :param excel_path: path to the existing Excel file
    :param sheet_name: target Excel sheet name
    :param column_letter: excel column letter to write into (e.g., "B")
    :param start_row: starting Excel row (default = 2, assuming headers in row 1).
    :param values: list of values to write into the specified column
    :return: Excel file updated with the new values in the specified
    """
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    for i, val in enumerate(values, start=start_row):
        cell = f"{column_letter}{i}"
        ws[cell] = val

    wb.save(excel_path)
    print(f"✅ Updated Excel file '{excel_path}' with values from in column {column_letter}.")

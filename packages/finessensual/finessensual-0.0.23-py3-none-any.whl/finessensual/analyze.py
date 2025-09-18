import argparse
import os, sys, glob, time
import subprocess
import re
from datetime import datetime
from pprint import pprint

from finessensual.parsers import create_projects_template, parse_projects, create_persons_template, parse_persons, parse_planning, parse_projectbudgets, parse_budgetboeking, parse_salaries, parse_fl_transactions, parse_ff_transactions, parse_booked_expenditures, parse_fixed_expenditures
from finessensual.indexers import create_salary_views, create_eqpoprovh_views
from finessensual.persons import Person
from finessensual.expenditures import ExpenditureType, ExpenditurePhase
from finessensual.budget import BudgetItem
from finessensual.helper import parse_month
from finessensual.openpyxl_helper import add_border

from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from .__about__ import __version__

# style sheet items
booked = PatternFill( start_color = "E0FFFF", end_color = "E0FFFF", fill_type = "solid" )
fixed = PatternFill( start_color = "E0FFE0", end_color = "E0FFE0", fill_type = "solid" )
planned = PatternFill( start_color = "FFE4B5", end_color = "FFE4B5", fill_type = "solid" )
overruled = PatternFill( start_color = "FFFFE0", end_color = "FFFFE0", fill_type = "solid" )

operational = PatternFill( start_color = "DDDDDD", end_color = "FFFFFF", fill_type = "solid" )
overhead = PatternFill( start_color = "EEEEEE", end_color = "EEEEEE", fill_type = "solid" )
salary = PatternFill( start_color = "CCCCCC", end_color = "CCCCCC", fill_type = "solid" )

colormap = { ExpenditurePhase.FIXED : fixed,
             ExpenditurePhase.BOOKED : booked,
             ExpenditurePhase.PLANNED : planned,
             ExpenditurePhase.OVERRULED: overruled }

boldfont = Font( bold = True )
alignleft   = Alignment( horizontal = "left" )
aligncenter = Alignment( horizontal = "center" )
alignright  = Alignment( horizontal = "right" )

currentyear = datetime.now().year;

files_toclean = []

def fatalerror( message: str ):
    sys.stderr.write( "Error: " + message )
    input( "\n\nPress any key to close this session..." )
    sys.exit()

def analyze( datadir: str,
             output : str ) -> bool:

    print( "\n".join( [ "=================================================",
                        f"finessensual v{__version__}",
                        "(C) 2025 Walter Daems - No Warranty Whatsoever",
                        "    This package is released under an MIT License",
                        "=================================================" ] ) )

    xlsxconvert: str
    match sys.platform:
        case 'win32':
            xlsxconvert = '/Program Files/LibreOffice/program/soffice.exe'
        case 'linux':
            xlsxconvert = '/usr/bin/soffice'
        case _:
            fatalerror( 'your OS is not yet supported, by me a machine and I will provide support' )
            
    if not ( os.path.isfile( xlsxconvert ) and os.access( xlsxconvert, os.X_OK ) ):
        fatalerror( "could not find a working installation of libreoffice. Please check your setup." )
            
    ##### 1. Parse projects
    projects = {}
    projects_filename = os.path.join( datadir, 'Projects.xlsx' )
    if not os.path.isfile( projects_filename ):
        print( f"- Creating empty {projects_filename}" )
        create_projects_template( projects_filename )
    else:
        print( f"- Parsing {projects_filename}" )
        workbook = load_workbook( projects_filename )
        sheet = workbook[ 'Projects' ]
        try:
            parse_projects( projects, sheet )
        except RuntimeError as e:
            fatalerror( f"could not parse projects.\n{e}" )
        except KeyError as e:
            fatalerror( f"could not parse projects (internal key error).\n{e}" )
            

    ##### 2. Parse persons
    persons = {}
    persons_filename = os.path.join( datadir, 'Persons.xlsx' )
    if not os.path.isfile( persons_filename ):
        print( f"- Creating empty {persons_filename}" )
        create_persons_template( persons_filename )
    else:
        print( f"- Parsing {persons_filename}" )
        workbook = load_workbook( persons_filename )
        sheet = workbook[ 'Persons' ]
        try:
            parse_persons( persons, sheet )
        except RuntimeError as e:
            fatalerror( f"could not parse persons.\n{e}" )
        except KeyError as e:
            fatalerror( f"could not parse projects (internal key error).\n{e}" )

    data = {}
    for t in ExpenditureType:
        data[t] = {}
        for p in ExpenditurePhase:
            data[t][p] = {}

    ##### 3. Parse planning
    planning = {}
    planning_filename = os.path.join( datadir, 'Planning.xlsx' )
    if not os.path.isfile( planning_filename ):
        print( f"- Create your own planning by isolating the Personnel tab in a file" )
    else:
        print( f"- Parsing {planning_filename}" )
        workbook = load_workbook( planning_filename )
        sheet = workbook[ 'Personnel' ]
        try:
            parse_planning( planning , sheet )
        except RuntimeError as e:
            fatalerror( f"could not parse planning.\n{e}" )
        except KeyError as e:
            fatalerror( f"could not parse projects (internal key error).\n{e}" )

        # now write the planned portion into the data:
        data[ExpenditureType.WEDDEN][ExpenditurePhase.PLANNED] = planning[ExpenditurePhase.PLANNED]

    errlog = open( 'finessensual.err', 'w' )
    
    ##### 4. Parse the budgets
    print( "- Parsing Budgets" )

    ublist = glob.glob( os.path.join( datadir, 'UB-*.xml' ) )
    if len( ublist ) == 0:
        fatalerror( f"could not find any 'UB-*.xml'" )
    for ub in ublist:
        print( f"  - {ub}" )
        ubdir, ubtarget = toxlsx( ub )
        files_toclean.append( ubtarget )
        retval = subprocess.run( [ xlsxconvert, '--convert-to', 'xlsx',
                                   '--outdir', f'{ubdir}', f'{ub}' ],
                                 stdout = errlog
                                )
        if not os.path.isfile( ubtarget ):
            fatalerror( f"the conversion of {ub} to {ubtarget} failed." )
        
        workbook = load_workbook( ubtarget )
        rapport = workbook[ 'Rapport' ]
        try:
            parse_projectbudgets( projects, rapport )
        except RuntimeError as e:
            fatalerror( f"could not parse project budget.\n{e}" )
        except KeyError as e:
            fatalerror( f"could not parse projects (internal key error).\n{e}" )

    bblist = glob.glob( os.path.join( datadir, 'Budgetboekingen-*.xls' ) )
    if len( bblist ) == 0:
        fatalerror( f"could not find any 'Budgetboekingen-*.xml'" )
    for bb in bblist:
        print( f"  - {bb}" )
        bbdir, bbtarget = toxlsx( bb )
        files_toclean.append( bbtarget )
        retval = subprocess.run( [ xlsxconvert, '--convert-to', 'xlsx',
                                   '--outdir', f'{bbdir}', f'{bb}' ],
                                 stdout = errlog
                                )
        if not os.path.isfile( bbtarget ):
            fatalerror( f"the conversion of {bb} to {bbtraget} failed." )
        workbook = load_workbook( bbtarget )
        sheet = workbook[ 'Data' ]
        try:
            parse_budgetboeking( projects, sheet )
        except RuntimeError as e:
            fatalerror( f"Error : could not parse budgetboeking.\n{e}" )
        except KeyError as e:
            fatalerror( f"could not parse projects (internal key error).\n{e}" )

        #print( "===== Budgetten =====" )
        #print( projects )


    ##### 5. Parse the transactions in separate categories
    print( "- Parsing Vlottende transacties" )
    vtlist = glob.glob( os.path.join( datadir, 'VBTransacties-*.xls' ) )
    if len( vtlist ) == 0:
        fatalerror( f"could not find any 'VBTransacties-*.xls'" )
    for vt in vtlist:
        print( f"  - {vt}" )
        vtdir, vttarget = toxlsx( vt )
        files_toclean.append( vttarget )
        retval = subprocess.run( [ xlsxconvert, '--convert-to', 'xlsx',
                                   '--outdir', f'{vtdir}', f'{vt}' ],
                                 stdout = errlog
                                )
        if not os.path.isfile( vttarget ):
            fatalerror( f"the conversion of {vt} to {vttarget} failed." )

        workbook = load_workbook( vttarget )
        sheet = workbook[ 'Data' ]
        parse_fl_transactions( data, sheet, persons, projects )
    
    print( "- Parsing Forfaitaire transacties" )
    fflist = glob.glob( os.path.join( datadir, 'Transacties-FF*.xls' ) )
    if  len( fflist ) == 0:
        fatalerror( f"could not find any 'Transacties-FF*.xls'" )
    for ff in fflist:
        print( f"  - {ff}" )
        ffdir, fftarget = toxlsx( ff )
        files_toclean.append( fftarget )
        retval = subprocess.run( [ xlsxconvert, '--convert-to', 'xlsx',
                                   '--outdir', f'{ffdir}', f'{ff}' ],
                                 stdout = errlog
                                )
        workbook = load_workbook( fftarget )
        sheet = workbook[ 'Data' ]
        parse_ff_transactions( data, sheet, persons, projects )
        
    ##### 6. Parse the salaries
    print( "- Parsing Lonen transacties" )
    ltlist = glob.glob( os.path.join( datadir, 'Lonen-transacties-*.xml' ) )
    if len( ltlist ) == 0:
        fatalerror( f"could not find any 'Lonen-transacties-*.xml'" )
    for lt in ltlist:
        print( f"  - {lt}" )
        ltdir, lttarget = toxlsx( lt )
        files_toclean.append( lttarget )
        retval = subprocess.run( [ xlsxconvert, '--convert-to', 'xlsx',
                                   '--outdir', f'{ltdir}', f'{lt}' ],
                                 stdout = errlog
                                )

        if not os.path.isfile( bbtarget ):
            fatalerror( f"the conversion of {lt} to {lttarget} failed." )

        workbook = load_workbook( lttarget )
        sheet = workbook[ 'Rapport' ]
        parse_salaries( data, sheet, persons )
        
    ##### 7. Parse the expenses
    print( "- Parsing Uitgaven" )
    uulist = glob.glob( os.path.join( datadir, 'UU-*.xml' ) )
    if len( uulist ) == 0:
        fatalerror( f"could not find any 'UU-*.xml'" )
    for uu in uulist:
        print( f"  - {uu}" )
        uudir, uutarget = toxlsx( uu )
        files_toclean.append( uutarget )
        retval = subprocess.run( [ xlsxconvert, '--convert-to', 'xlsx',
                                   '--outdir', f'{uudir}', f'{uu}' ],
                                 stdout = errlog
                                )

        if not os.path.isfile( bbtarget ):
            fatalerror( f"the conversion of {uu} to {uutarget} failed." )

        workbook = load_workbook( uutarget )
        sheet = workbook[ 'Rapport' ]
        parse_booked_expenditures( data, sheet, persons )
        
    #print( "===== Wedden =========" )
    #print( data[ExpenditureType.WEDDEN][ExpenditurePhase.BOOKED] )
    #print( "===== Operations =====" )
    #print( data[ExpenditureType.WERKING][ExpenditurePhase.BOOKED] )
    #print( "===== Overhead =======" )
    #print( data[ExpenditureType.OVERHEAD][ExpenditurePhase.BOOKED] )
    
    ##### 8. Parse the fixations
    print( "- Parsing Vastleggingen" )
    uvlist = glob.glob( os.path.join( datadir, 'UV-*.xml' ) )
    if len( uvlist ) == 0:
        fatalerror( f"could not find any 'UV-*.xml'" )
    for uv in uvlist:
        print( f"  - {uv}" )
        uvdir, uvtarget = toxlsx( uv )
        files_toclean.append( uvtarget )
        retval = subprocess.run( [ xlsxconvert, '--convert-to', 'xlsx',
                                   '--outdir', f'{uvdir}', f'{uv}' ],
                                 stdout = errlog
                                )

        if not os.path.isfile( bbtarget ):
            fatalerror( f"the conversion of {uv} to {uvtarget} failed." )

        workbook = load_workbook( uvtarget )
        sheet = workbook[ 'Rapport' ]
        parse_fixed_expenditures( data, sheet, persons )

    errlog.close()
    
    #print( "===== Wedden =========" )
    #print( data[ExpenditureType.WEDDEN][ExpenditurePhase.FIXED] )
    #print( "===== Operations =====" )
    #print( data[ExpenditureType.WERKING][ExpenditurePhase.FIXED] )
    #print( "===== Overhead =======" )
    #print( data[ExpenditureType.OVERHEAD][ExpenditurePhase.FIXED] )

    ##### 9. Create the views
    views = {}
    for t in ExpenditureType:
        views[t] = {}
        for p in ExpenditurePhase:
            views[t][p] = {}
    
    create_salary_views( data[ExpenditureType.WEDDEN][ExpenditurePhase.BOOKED],
                         views[ExpenditureType.WEDDEN][ExpenditurePhase.BOOKED] )
    create_salary_views( data[ExpenditureType.WEDDEN][ExpenditurePhase.FIXED],
                         views[ExpenditureType.WEDDEN][ExpenditurePhase.FIXED] )
    create_salary_views( data[ExpenditureType.WEDDEN][ExpenditurePhase.PLANNED],
                         views[ExpenditureType.WEDDEN][ExpenditurePhase.PLANNED] )

    #for p in ExpenditurePhase:
        #print( "==== Salary views - " + str( p ) )
        #pprint( views[ExpenditureType.WEDDEN][p] )
    
    # print( "##########################" )
    # print( views[ExpenditureType.WEDDEN][ExpenditurePhase.FIXED] )
    # print( "##########################" )
    
    create_eqpoprovh_views( data[ExpenditureType.UITRUSTING][ExpenditurePhase.BOOKED],
                            views[ExpenditureType.UITRUSTING][ExpenditurePhase.BOOKED] )
    create_eqpoprovh_views( data[ExpenditureType.UITRUSTING][ExpenditurePhase.FIXED],
                            views[ExpenditureType.UITRUSTING][ExpenditurePhase.FIXED] )
    
    create_eqpoprovh_views( data[ExpenditureType.WERKING][ExpenditurePhase.BOOKED],
                            views[ExpenditureType.WERKING][ExpenditurePhase.BOOKED] )
    create_eqpoprovh_views( data[ExpenditureType.WERKING][ExpenditurePhase.FIXED],
                            views[ExpenditureType.WERKING][ExpenditurePhase.FIXED] )
    
    create_eqpoprovh_views( data[ExpenditureType.OVERHEAD][ExpenditurePhase.BOOKED],
                            views[ExpenditureType.OVERHEAD][ExpenditurePhase.BOOKED] )
    create_eqpoprovh_views( data[ExpenditureType.OVERHEAD][ExpenditurePhase.FIXED],
                            views[ExpenditureType.OVERHEAD][ExpenditurePhase.FIXED] )
    

    ##### 10. Compose the output file
    print( f"- Creating output file {output}" )

    finesse = Workbook()
    finesse.remove(finesse.active)

    ##### 10.A Information sheet
    print( f"  - Information tab page" )

    information = finesse.create_sheet( 'Information' )
    for c in range( 1, 15 ):
        information.column_dimensions[ get_column_letter( c ) ].width = 20
    
    labelcol = information.column_dimensions['A']
    datacol = information.column_dimensions['B']
    
    information.append( [ 'Data:', 'Essential financial information for research groups of University of Antwerp' ] )
    information.append( [ 'Generated on:', datetime.now() ] )
    information.append( [ 'Generated by:', f'finessensual (C) {currentyear} Walter Daems' ] )
    information.append( [ 'Disclaimer:', 'This data has been generated by a program under MIT license' ] )
    information.append( [ '', 'No Warranty Whatsoever' ] )
    
    labelcol.font = boldfont
    labelcol.alignment = alignright
    datacol.alignment = alignleft
    information.sheet_view.zoomScale = 80

    #############################
    ###### 10.B Personnel sheet
    print( f"  - Personnel tab page" )
    sheet = finesse.create_sheet( 'Personnel' )
    
    sheet.column_dimensions[ get_column_letter( 1 ) ].width = 12
    
    sheet['A1'] = 'Personnel'
    sheet['A5'] = 'Month'
    
    # first need get an overall overview: extreme dates and pdata
    monthlist = {}
    personsizelist = {}
    start_month = f"{currentyear}-01"
    pdata = {} # pdata[month][personnr][projectnr] = [phase, amount]
    for phase in ( ExpenditurePhase.PLANNED, ExpenditurePhase.FIXED, ExpenditurePhase.BOOKED ):
        currentview = views[ExpenditureType.WEDDEN][phase]['by_date']
        currentdata = data[ExpenditureType.WEDDEN][phase]
        for month, persondict in currentview.items():
            if month == 'None' or month < start_month:
                continue
            if month not in monthlist:
                monthlist[month] = 1
                if month not in pdata:
                    pdata[month] = {}
            for personnr, projectdict in persondict.items():
                if personnr not in pdata[month]:
                    pdata[month][personnr] = {}
                if personnr not in personsizelist:
                    personsizelist[personnr] = len( projectdict.keys() )
                else:
                    personsizelist[personnr] = max( personsizelist[personnr], len( projectdict.keys() ) )
                for projectnr, key in projectdict.items():
                    pdata[month][personnr][projectnr] = [phase, currentdata[key] ]
                    # this will overwrite the less important phases with the more important phases
    
    monthlist = { key: index for index, key in enumerate( sorted( monthlist.keys() ) ) }
    
    table_row = 6
    
    for month, i in monthlist.items():
        sheet.cell( table_row + i, 1, month )
    
        
    col = 2
    for personnr, size in sorted( personsizelist.items() ):
        cell = sheet.cell( table_row - 3, col, personnr )
        cell.font = boldfont;
        cell.alignment = Alignment( horizontal = "center" )
        sheet.merge_cells( start_row = table_row - 3, end_row = table_row -3,
                           start_column = col, end_column = col + 3 * size -1 )
        cell = sheet.cell( table_row - 2, col, persons[personnr].name )
        cell.font = boldfont;
        cell.alignment = Alignment( horizontal = "center" )
        sheet.merge_cells( start_row = table_row - 2, end_row = table_row -2,
                           start_column = col, end_column = col + 3 * size -1 )
        for i in range( 0, size ):
            colindex = col + i * 3
            sheet.column_dimensions[ get_column_letter( colindex ) ].width = 10
            sheet.column_dimensions[ get_column_letter( colindex + 1 ) ].width = 3
            sheet.column_dimensions[ get_column_letter( colindex + 2 ) ].width = 14
            cell = sheet.cell( table_row -1, colindex , 'Project' )
            cell.font = boldfont;
            cell.alignment = alignright
            cell = sheet.cell( table_row -1, colindex + 1, 'Ph' )
            cell.font = boldfont;
            cell.alignment = aligncenter
            cell = sheet.cell( table_row -1, colindex + 2, 'Amount   ' )
            cell.font = boldfont;
            cell.alignment = alignright
    
        for month, i in monthlist.items():
            if personnr in pdata[month]:
                j = 0
                for projectnr, array in pdata[month][personnr].items():
                    cell = sheet.cell( table_row + i, col + j,
                                       '=HYPERLINK("{}", "{}")'.format( "#{}".format( projects[str(projectnr)].sheetname() ),
                                                                        projectnr ) )
                    cell.alignment = alignright
                    cell.fill = colormap[array[0]]
                    cell = sheet.cell( table_row + i, col + j + 1, array[0].name[0] )
                    cell.alignment = aligncenter
                    cell = sheet.cell( table_row + i, col + j + 2, array[1].totalcost() )
                    cell.comment = Comment( array[1].overview(), 'Summary' )
                    cell.fill = colormap[array[0]]
                    cell.number_format = '#,##0.00 €   '
                    j += 3
        add_border( sheet, min_row = 3, max_row = table_row + len( monthlist.keys() ) - 1 ,
                    min_col = col, max_col = col + 3*size - 1 )
        col += 3 * size
    
    legendrow = sheet.max_row + 2
    row = legendrow
    cell = sheet.cell( row, 4, "Project legend" )
    cell.font = boldfont
    
    for nr, project in projects.items():
        row += 1
        sheet.cell( row, 4, nr )
        sheet.cell( row, 5, project.name )
    
    row = legendrow
    cell = sheet.cell( row, 9, "Phase legend" )
    cell.font = boldfont
    for j, phase in enumerate( ExpenditurePhase ):
        row += 1
        sheet.column_dimensions[ get_column_letter( colindex ) ].width = 14
        cell = sheet.cell( row, 10, phase.name )
        cell.alignment = Alignment( horizontal = "center" )
        cell.fill = colormap[ phase ]
        
    for i in range(1,6):
        labelrow = sheet.row_dimensions[i]
        labelrow.font = boldfont

    sheet.freeze_panes = sheet[ 'B6' ]
    sheet.sheet_view.zoomScale = 80

    ###################################
    ###### 10.C Expense budgets sheet
    print( f"  - Expense budgets tab page" )
    ebsheet = finesse.create_sheet( 'ExpBudgets' )
    
    ebsheet.column_dimensions[ get_column_letter( 1 ) ].width = 20
    
    cell = ebsheet.cell( 1, 1, 'Expense budgets (saldi)' )
    cell.font = boldfont

    cell = ebsheet.cell( 3, 1, 'Project' )
    cell.font = boldfont 
    ebsheet.merge_cells ( start_row = 3, end_row = 4,
                          start_column = 1, end_column = 1 )
   
    mycol = 2
    for label in [ 'Equipment', 'Operational' ]:
        cell = ebsheet.cell( 3, mycol, label )
        cell.font = boldfont
        ebsheet.merge_cells ( start_row = 3, end_row = 3,
                              start_column = mycol, end_column = mycol+4 )
        mycol += 5

    mycol = 2
    for j in range(1,3):
        for label in [ 'Planned', 'Fixed', 'Booked', 'Overruled', 'Remaining' ]:
            ebsheet.column_dimensions[ get_column_letter( mycol ) ].width = 16
            cell = ebsheet.cell( 4, mycol, label + "   " )
            cell.alignment = alignright
            cell.font = boldfont
            mycol += 1

    ebrow = 4
    ###########################
    ###### 10.D Project tabs
    print( f"  - Projects tab pages" )
    for key, project in projects.items():

        print( f"    - {key} - {project.sheetname()}" )
        
        # create new sheet
        sheet = finesse.create_sheet( project.sheetname() )
    
        # global column planning
        columns = { ExpenditureType.WEDDEN: {'start': 14, 'end': 18 },
                    ExpenditureType.WERKING: {'start': 6, 'end': 10 },
                    ExpenditureType.UITRUSTING: {'start': 1, 'end': 5 },
                    ExpenditureType.OVERHEAD: {'start': 11,'end': 13 },
                   }
        
        # set the columns to a decent size
        for c in range( 1, 30 ):
            sheet.column_dimensions[ get_column_letter( c ) ].width = 18
    
        # create info data
        colA = sheet.column_dimensions['A']
        colA.font = boldfont
        colB = sheet.column_dimensions['B']
        colC = sheet.column_dimensions['C']
        row = 1
        sheet.cell( row, 1, project.blurb() )
        row += 2
        for id in [ [ 'Promotor:', project.promotor ],
                    [ 'Copromotor:', project.copromotor ],
                    [ 'Budgetcode:', project.budgetcode() ],
                    [ 'Source:', project.source ],
                    [ 'Begin date:', project.begindate ],
                    [ 'End date:', project.enddate ],
                   ]:
            cell = sheet.cell( row, 1, id[0] )
            cell.alignment = alignright
            cell.font = boldfont
            sheet.cell( row, 2, id[1] )
            sheet.merge_cells ( start_row = row, end_row = row, start_column = 2, end_column = 3 )
            row += 1
        add_border( sheet,
                    min_row = 3, max_row = row - 1,
                    min_col = 1, max_col = 3 )
    
            
        # create budget info
        budget_row = 3
        start_col = 4
        oracle_col = start_col + 2
        cellrange = sheet.merge_cells( start_row = budget_row - 1, end_row = budget_row - 1,
                                       start_column = oracle_col, end_column = oracle_col + 3 )
        cell = sheet.cell( budget_row - 1, oracle_col, "From Oracle" )
        cell.font = boldfont
        add_border( sheet,
                    min_row = budget_row, max_row = budget_row + 5,
                    min_col = start_col, max_col = start_col + 1 )
        add_border( sheet,
                    min_row = budget_row, max_row = budget_row + 5,
                    min_col = oracle_col, max_col = oracle_col + 3 )
        for j, v in enumerate( [ 'Category', *[ key + "   " for key in BudgetItem.oraclelabels() ] ] ):
            cell = sheet.cell( budget_row, start_col + j, v )
            cell.alignment = alignright
            cell.font = boldfont
        if project.issimple():
            for j, v in enumerate( ['OVERALL', *project.budget.overview() ] ):
                cell = sheet.cell( budget_row + 1, start_col + j, v )
                cell.number_format = '#,##0.00 €   '
                cell.alignment = alignright
        else:
            for i, b in enumerate( ExpenditureType ):
                if project.budget[b]:
                    for j, v in enumerate( [b.name, *project.budget[b].overview() ] ):
                        cell = sheet.cell( budget_row + i + 1, start_col + j, v )
                        cell.number_format = '#,##0.00 €   '
                        cell.alignment = alignright
    
        for j, phase in enumerate( ExpenditurePhase ):
            cell = sheet.cell( j+4, 16, phase.name )
            cell.alignment = Alignment( horizontal = "center" )
            cell.fill = colormap[ phase ] 
            
        # create salary info
        detail_row = sheet.max_row + 2
        start_row = detail_row
        salary_col = columns[ExpenditureType.WEDDEN]['start']
        cell = sheet.cell( start_row, salary_col, 'Salaries' )
        cell.font = boldfont
    
        cell = sheet.cell( start_row + 1, salary_col, 'Per month' )
        cell.alignment = alignright
        cell.font = boldfont
        
        ##############
        # we will first build the overview in memory and assess the size of the table (in columns)
    
        # initialize the detail totals
        totals = {}
        for t in ExpenditureType:
            totals[t] = {}
            for p in ExpenditurePhase:
                totals[t][p] = 0.0
    
        # column and row management
        person_column  = {}  # dict from personnr to column number
        person_counter = 1  # will govern the columns (we start from 2 as column 1 will contain the date
        date_row       = {}  # dict from date tot row number, will first be all ones
        
        # let's provide a structure to keep track of the salary costs
        # salarycosts : dict[ date, dict [ person, [ExpenditureType, salarycost ] ] ]
        salarycosts = {}
        # we will treat the phases in order from uncertain to certain to overwrite old plannings
        for phase in ( ExpenditurePhase.PLANNED, ExpenditurePhase.FIXED, ExpenditurePhase.BOOKED ):
            currentview = views[ExpenditureType.WEDDEN][phase]['by_project']
            currentdata = data[ExpenditureType.WEDDEN][phase]
            if project.number in currentview:
                for date, personnr_dict in sorted( currentview[project.number].items() ):
                    ### If the project is 'vlottend', then only the current year needs to be considered
                    if re.search( r"^[AWR]", project.number ):
                        if int(date[0:4]) < currentyear:
                            continue;
                    if date not in date_row:
                        date_row[date] = 1 # we will arrange the row number later
                        salarycosts[date] = {}
                    for personnr, keylist in sorted( personnr_dict.items() ):
                        if personnr not in person_column:
                            person_column[personnr] = columns[ExpenditureType.WEDDEN]['start'] + person_counter
                            person_counter += 1
                        assert len(keylist) == 1, f"Internal data inconsistency: double booking of same person {personnr} in same month {date} on same project {project.number}"
                        salarycosts[date][personnr] = [ phase, currentdata[keylist[0]] ]

        # now we rearrange the dates in order
        date_row = { key: index + start_row + 2 for index, key in enumerate( sorted( date_row.keys() ) ) }
    

        for person, col in person_column.items():
            cell = sheet.cell( start_row + 1, col, persons[person].name )
            cell.font = boldfont
            cell.alignment = alignright
        for date, personlist in salarycosts.items():
            cell = sheet.cell( date_row[date], salary_col, date )
            cell.alignment = alignright
            for person, phaseandcost in personlist.items():
                (phase, cost) = phaseandcost
                totals[ExpenditureType.WEDDEN][phase] += cost.totalcost()
                cell = sheet.cell( date_row[date], person_column[person], cost.totalcost() )
                cell.alignment = alignright
                cell.number_format = '#,##0.00 €   '
                cell.fill = colormap[ phase ]
                cell.comment = Comment( cost.overview(), 'Summary' )
                # print( f"Writing {cost} for {person} in {phase} on {date} on project {key}\n" )

        columns[ExpenditureType.WEDDEN]['end'] = columns[ExpenditureType.WEDDEN]['start'] + person_counter - 1
        sheet.merge_cells( start_row = start_row, end_row = start_row,
                           start_column = columns[ExpenditureType.WEDDEN]['start'],
                           end_column = columns[ExpenditureType.WEDDEN]['end'] )
    
        #########################################################
        # create equipment costs info
        start_row = detail_row
        cell = sheet.cell( start_row, columns[ExpenditureType.UITRUSTING]['start'], 'Equipment' )
        cell.font = boldfont
        sheet.merge_cells( start_row = start_row, end_row = start_row,
                           start_column = columns[ExpenditureType.UITRUSTING]['start'],
                           end_column = columns[ExpenditureType.UITRUSTING]['start'] + 4 )
    
        for i, header in enumerate( [ 'Per date', 'Batchnr', 'Amount', 'Description' ] ):
            cell = sheet.cell( start_row + 1, i + columns[ExpenditureType.UITRUSTING]['start'], header )
            cell.font = boldfont
            cell.alignment = alignright
            cell.alignment = Alignment( horizontal = "left" )
            sheet.merge_cells( start_row = start_row + 1, end_row = start_row + 1,
                               start_column = columns[ExpenditureType.UITRUSTING]['start'] + 3,
                               end_column = columns[ExpenditureType.UITRUSTING]['start'] + 4 )
            
        # let's provide a structure to keep track of the operational costs
        # operationalcosts : dict[ date, [ExpenditureType, operational ] ] ]
        equipmentcosts = {}
        # we will treat the phases in order from uncertain to certain to overwrite old plannings
        for phase in ( ExpenditurePhase.FIXED, ExpenditurePhase.BOOKED ):
            currentview = views[ExpenditureType.UITRUSTING][phase]['by_project']
            currentdata = data[ExpenditureType.UITRUSTING][phase]
            if project.number in currentview:
                for date, costlist in sorted( currentview[project.number].items() ):
                    if date not in equipmentcosts:
                        equipmentcosts[date] = []
                    for cost in costlist:
                        equipmentcosts[date].append( [ phase, currentdata[cost] ] )
    
        row = start_row + 2
        for date, phasecosts in sorted( equipmentcosts.items () ):
            for phasecost in phasecosts:
                (phase, cost) = phasecost
                totals[ExpenditureType.UITRUSTING][phase] += cost.totalcost()
                cell = sheet.cell(row, columns[ExpenditureType.UITRUSTING]['start'], cost.date )
                cell.alignment = alignright
                cell = sheet.cell(row, columns[ExpenditureType.UITRUSTING]['start'] + 1, cost.batchnr )
                cell.alignment = alignright
                cell = sheet.cell(row, columns[ExpenditureType.UITRUSTING]['start'] + 2, cost.totalcost() )
                cell.alignment = alignright
                cell.fill = colormap[ phase ]
                cell.number_format = '#,##0.00 €   '
                cell.comment = Comment( cost.overview(), 'Summary' )
                cell = sheet.cell(row, columns[ExpenditureType.UITRUSTING]['start'] + 3, cost.nature )
                cell.alignment = Alignment( horizontal = "left" )
                sheet.merge_cells( start_row = row, end_row = row, start_column = columns[ExpenditureType.UITRUSTING]['start'] + 3, end_column = columns[ExpenditureType.UITRUSTING]['start'] + 4 )
                row += 1
    
        ##############################################
        # create operational costs info
        start_row = detail_row
        cell = sheet.cell( start_row, columns[ExpenditureType.WERKING]['start'], 'Operational' )
        cell.font = boldfont
        sheet.merge_cells( start_row = start_row, end_row = start_row,
                           start_column = columns[ExpenditureType.WERKING]['start'],
                           end_column = columns[ExpenditureType.WERKING]['start'] + 4 )
    
        for i, header in enumerate( [ 'Per date', 'Batchnr', 'Amount', 'Description' ] ):
            cell = sheet.cell( start_row + 1, i + columns[ExpenditureType.WERKING]['start'], header )
            cell.font = boldfont
            cell.alignment = alignright
            cell.alignment = Alignment( horizontal = "left" )
            sheet.merge_cells( start_row = start_row + 1, end_row = start_row + 1,
                               start_column = columns[ExpenditureType.WERKING]['start'] + 3,
                               end_column = columns[ExpenditureType.WERKING]['start'] + 4 )
            
        # let's provide a structure to keep track of the operational costs
        # operationalcosts : dict[ date, [ExpenditureType, operational ] ] ]
        operationalcosts = {}
        # we will treat the phases in order from uncertain to certain to overwrite old plannings
        for phase in ( ExpenditurePhase.FIXED, ExpenditurePhase.BOOKED ):
            currentview = views[ExpenditureType.WERKING][phase]['by_project']
            currentdata = data[ExpenditureType.WERKING][phase]
            if project.number in currentview:
                for date, costlist in sorted( currentview[project.number].items() ):
                    if date not in operationalcosts:
                        operationalcosts[date] = []
                    for cost in costlist:
                        operationalcosts[date].append( [ phase, currentdata[cost] ] )
    
        row = start_row + 2
        for date, phasecosts in sorted( operationalcosts.items () ):
            for phasecost in phasecosts:
                (phase, cost) = phasecost
                totals[ExpenditureType.WERKING][phase] += cost.totalcost()
                cell = sheet.cell(row, columns[ExpenditureType.WERKING]['start'], cost.date )
                cell.alignment = alignright
                cell = sheet.cell(row, columns[ExpenditureType.WERKING]['start'] + 1, cost.batchnr )
                cell.alignment = alignright
                cell = sheet.cell(row, columns[ExpenditureType.WERKING]['start'] + 2, cost.totalcost() )
                cell.alignment = alignright
                cell.fill = colormap[ phase ]
                cell.number_format = '#,##0.00 €   '
                cell.comment = Comment( cost.overview(), 'Summary' )
                cell = sheet.cell(row, columns[ExpenditureType.WERKING]['start'] + 3, cost.nature )
                cell.alignment = Alignment( horizontal = "left" )
                sheet.merge_cells( start_row = row, end_row = row, start_column = columns[ExpenditureType.WERKING]['start'] + 3, end_column = columns[ExpenditureType.WERKING]['start'] + 4 )
                row += 1
    
        # create overhead info
        start_row = detail_row
        cell = sheet.cell( start_row, columns[ExpenditureType.OVERHEAD]['start'], 'Overhead' )
        cell.font = boldfont
        sheet.merge_cells( start_row = start_row, end_row = start_row,
                           start_column = columns[ExpenditureType.OVERHEAD]['start'],
                           end_column = columns[ExpenditureType.OVERHEAD]['start'] + 2 )
    
        for i, header in enumerate( [ 'Per date', 'Batchnr', 'Amount' ] ):
            cell = sheet.cell( start_row + 1, i + columns[ExpenditureType.OVERHEAD]['start'], header )
            cell.font = boldfont
            cell.alignment = alignright
            
        # let's provide a structure to keep track of the overhead costs
        # overheadcosts : dict[ date, [ExpenditureType, overhead ] ] ]
        overheadcosts = {}
        # we will treat the phases in order from uncertain to certain to overwrite old plannings
        for phase in ( ExpenditurePhase.FIXED, ExpenditurePhase.BOOKED ):
            currentview = views[ExpenditureType.OVERHEAD][phase]['by_project']
            currentdata = data[ExpenditureType.OVERHEAD][phase]
            if project.number in currentview:
                for date, costlist in sorted( currentview[project.number].items() ):
                    if date not in overheadcosts:
                        overheadcosts[date] = []
                    for cost in costlist:
                        overheadcosts[date].append( [ phase, currentdata[cost] ] )
    
        row = start_row + 2
        for date, phasecosts in sorted( overheadcosts.items () ):
            for phasecost in phasecosts:
                (phase, cost) = phasecost
                totals[ExpenditureType.OVERHEAD][phase] += cost.totalcost()
                cell = sheet.cell(row, columns[ExpenditureType.OVERHEAD]['start'], cost.date )
                cell.alignment = alignright
                cell = sheet.cell(row, columns[ExpenditureType.OVERHEAD]['start'] + 1, cost.batchnr )
                cell.alignment = alignright
                cell = sheet.cell(row, columns[ExpenditureType.OVERHEAD]['start'] + 2, cost.totalcost() )
                cell.alignment = alignright
                cell.fill = colormap[ phase ]
                cell.number_format = '#,##0.00 €   '
                cell.comment = Comment( cost.overview(), 'Summary' )
                row += 1
                
        for t in ExpenditureType:
            if t in columns:
                while columns[t]['end'] - columns[t]['start'] < 2:
                    columns[t]['end'] += 1
                add_border( sheet,
                            min_row = detail_row, max_row = sheet.max_row,
                            min_col = columns[t]['start'],
                            max_col = columns[t]['end'] )
                    
        #######################################################################
        # Make the detail totals
        total_row = sheet.max_row + 1
        for t in ExpenditureType:
            if t in columns:
                row = total_row
                cell = sheet.cell( row, columns[t]['start'], 'Totals' )
                cell.font = boldfont
                cell.alignment = Alignment( "right" )
                row  += 1
                for p in ExpenditurePhase:
                    cell = sheet.cell( row, columns[t]['start'] + 1, p.name )
                    cell.alignment = alignright
                    cell = sheet.cell( row, columns[t]['start'] + 2, totals[t][p] )
                    cell.number_format = '#,##0.00 €   '
                    row += 1
                add_border( sheet,
                            min_row = total_row, max_row = row - 1,
                            min_col = columns[t]['start'], max_col = columns[t]['end'] )
    
        ##################################################################
        # Write summaries 'from planning' in the top=-right
        # --> in the mean time we also write the data in ExpBudgets

        # write project name on ebsheet
        ebrow += 1
        ebsheet.cell( ebrow, 1, project.sheetname() )
        
        planning_col = oracle_col + 4
        cellrange = sheet.merge_cells( start_row = budget_row - 1, end_row = budget_row - 1,
                                       start_column = planning_col, end_column = planning_col + 4 )
        cell = sheet.cell( budget_row - 1, planning_col, "From Planning" )
        cell.font = boldfont
        add_border( sheet,
                    min_row = budget_row, max_row = budget_row + 5,
                    min_col = planning_col, max_col = planning_col + 4 )
        for j, v in enumerate( [ key + "   " for key in BudgetItem.planninglabels() ] ):
            cell = sheet.cell( budget_row, planning_col + j, v )
            cell.alignment = alignright
            cell.font = boldfont
        if project.issimple():
            mytotals = {}
            subtotal = 0.0
            for j, phase in enumerate( ExpenditurePhase ):
                mytotals[phase] = 0.0
                for i, type in enumerate( ExpenditureType ):
                    mytotals[phase] += totals[type][phase]
                    subtotal += totals[type][phase]
                    cell = sheet.cell( budget_row + 1, planning_col + j, mytotals[phase] )
                    cell.number_format = '#,##0.00 €   '
                    cell.alignment = alignright
                    cell = sheet.cell( budget_row + 1, planning_col+4, project.budget.total - subtotal )
                    cell.number_format = '#,##0.00 €   '
                    cell.alignment = alignright

            for j, phase in enumerate( ExpenditurePhase ):
                cell = ebsheet.cell( ebrow, 7+j, mytotals[phase] )
                cell.number_format = '#,##0.00 €   '
            cell = ebsheet.cell( ebrow, 11, project.budget.total - subtotal )
            cell.number_format = '#,##0.00 €   '

        else:
            subtotals = {}
            for i, type in enumerate( ExpenditureType ):
                if type in columns:
                    subtotals[type] = 0
                    for j, phase in enumerate( ExpenditurePhase ):
                        subtotals[type] += totals[type][phase]
                        cell = sheet.cell( budget_row + i + 1, planning_col + j, totals[type][phase] )
                        cell.number_format = '#,##0.00 €   '
                        cell.alignment = alignright
                        cell = sheet.cell( budget_row + i + 1, planning_col+4, project.budget[type].total - subtotals[type] )
                        cell.number_format = '#,##0.00 €   '
                        cell.alignment = alignright
                        
            # complete writing the ExpBudgets
            for j, phase in enumerate( ExpenditurePhase ):
                cell = ebsheet.cell( ebrow, 2+j, totals[ExpenditureType.UITRUSTING][phase] )
                cell.number_format = '#,##0.00 €   '
            cell = ebsheet.cell( ebrow, 6, project.budget[ExpenditureType.UITRUSTING].total
                                 - subtotals[ExpenditureType.UITRUSTING] )
            cell.number_format = '#,##0.00 €   '
            
            for j, phase in enumerate( ExpenditurePhase ):
                cell = ebsheet.cell( ebrow, 7+j, totals[ExpenditureType.WERKING][phase] )
                cell.number_format = '#,##0.00 €   '
            cell = ebsheet.cell( ebrow, 11, project.budget[ExpenditureType.WERKING].total
                                 - subtotals[ExpenditureType.WERKING] )
            cell.number_format = '#,##0.00 €   '
                 
        sheet.freeze_panes = sheet[ 'A12' ]
        sheet.sheet_view.zoomScale = 80
        
    print( f"  - Saving it all" )
        
    finesse.save( output )

    # generate error such that hatch test produces output
    # print( f"{abababa}" );

    print( f"  - Cleaning temporary files" )
    for f in files_toclean:
        os.remove( f )

    print( "Done" )
        
    return True 
    
    # on persons tab:
    # for every person

def main() -> bool:
    parser = argparse.ArgumentParser( prog = "finessensual",
                                      description = "Generation of essential financial information for Research Teams of the University of Antwerp",
                                      epilog = "Missing a feature? Contact walter.daems@uantwerpen.be",
                                     )
    parser.add_argument( '-v', '--version',
                         action='store_true',
                         help = 'display the current version number' )
    parser.add_argument( '-o', '--output',
                         type = str,
                         default = 'financial-overview.xlsx',
                         help = 'filename of the excel to write the output to (default = output.xlsx)' )
    parser.add_argument( 'data',
                         nargs='?',
                         type = str,
                         help = 'name of the directory that contains the oracle excel data files (mandatory unless [-h] or [-v])' )

    args = parser.parse_args()

    if args.version:
        sys.stderr.write( f"finessensual v{__version__}\nCopyright (C) 2025 Walter Daems\nThis package is released under an MIT License\n" )
        sys.exit(0)

    if not args.data:
        parser.print_help();
        sys.exit(1)
        
    if not os.path.isdir( args.data ):
        fatalerror( f"could not locate data directory '{args.data}'" )
    output = os.path.join( args.data, args.output )
    if os.path.isdir( output ):
        fatalerror( f"output file '{output}' is obstructed by a directory with the same name" )
    if os.path.isfile( output ) and not os.access( output, os.W_OK ):
        fatalerror( f"file '{output}' is not writable" )

    try:
        analyze( datadir = args.data,
                 output  = output )
    except FileNotFoundError as fnfe:
        fatalerror( f"could not find data source file '{fnfe.filename}'" )
        
    time.sleep(2)

    return True;

def toxlsx( path: str ) -> str:
    retval = os.path.splitext( path )[0] + ".xlsx"
    if path == retval:
        fatalerror( "dangerous internal error with risk for data loss - aborting!\n" +
                    "Contact Walter Daems to report this." );
    return os.path.dirname( path ), retval
    
    
    

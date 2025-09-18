from finessensual.budget import BudgetItem
from finessensual.costs import SalaryCost, OperationCost, EquipmentCost, OverheadCost
from finessensual.expenditures import ExpenditureType, ExpenditurePhase, expenditure_phase
from finessensual.persons import Person, parse_person
from finessensual.projects import Project
from finessensual.helper import parse_day, parse_month, parse_noyear

import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import re
import sys

def create_projects_template( path: str ):
    projects = Workbook()
    projects.remove(projects.active)

    sheet = projects.create_sheet( 'Projects' )

    labels = [ 'Nr', 'Name', 'ExpenditureTypeless', 'Sheetname (max31)',
               'Promotor', 'Copromotor', 'Afdeling', 'Eenheid', 'Source',
               'Begindate', 'Enddate' ]
    for row in sheet.iter_rows( min_col = 0, max_col = len(labels),
                                min_row = 1, max_row = 1 ):
        for colnr, cell in enumerate( row ):
            sheet.column_dimensions[ get_column_letter( colnr + 1 ) ].width = 20
            cell.value = labels[colnr]
            cell.font = Font( name = 'Calibri', bold = True )

    projects.save( path )
    

def parse_projects( projects: dict,
                    sheet: openpyxl.workbook.workbook.Workbook ):
    # Get the label header line in the 'rapportProjects' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 1, max_row = 1, values_only = True ) ) )
    labelindex = {label: index for index, label in enumerate( labels ) }

    for r in sheet.iter_rows( min_row = 2, max_row = sheet.max_row, values_only = True ):
        project = str( r[labelindex['Nr']] )
        projects[project] = Project( r[labelindex['Name']],
                                     project,
                                     r[labelindex['ExpenditureTypeless']],
                                     r[labelindex['Sheetname (max31)']],
                                     r[labelindex['Promotor']],
                                     r[labelindex['Copromotor']],
                                     r[labelindex['Afdeling']],
                                     r[labelindex['Eenheid']],
                                     r[labelindex['Source']],
                                     r[labelindex['Begindate']],
                                     r[labelindex['Enddate']]
                                    )
        # initialize all budgets to zero
        bi = BudgetItem( 0, 0, 0 )
        if r[labelindex['ExpenditureTypeless']] == 'Y':
            projects[project].budget = bi
        else:
            for i in ExpenditureType:
                projects[project].budget[i] = bi

def create_persons_template( path: str ):

    persons = Workbook()
    persons.remove(persons.active)

    sheet = persons.create_sheet( 'Persons' )

    labels = [ 'Nr', 'Name', 'Start', 'End' ]
    for row in sheet.iter_rows( min_col = 0, max_col=len( labels ),
                                min_row = 1, max_row = 1 ):
        for colnr, cell in enumerate( row ):
            sheet.column_dimensions[ get_column_letter( colnr + 1 ) ].width = 20
            cell.value = labels[colnr]
            cell.font = Font( name = 'Calibri', bold = True )

    persons.save( path )
                
def parse_persons( persons: dict[ str, Person ],
                   sheet: openpyxl.workbook.workbook.Workbook ):
    # Get the label header line in the 'rapportProjects' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 1, max_row = 1, values_only = True ) ) )
    labelindex = {label: index for index, label in enumerate( labels ) }

    for r in sheet.iter_rows( min_row = 2, max_row = sheet.max_row, values_only = True ):
        personnr = "{:05d}".format( r[labelindex['Nr']] )
        name     = r[labelindex['Name']]
        start    = r[labelindex['Start']]
        end      = r[labelindex['End']]
        persons[personnr] = Person( personnr, name, start, end )
        
def parse_planning( planning: dict,
                    sheet: openpyxl.workbook.workbook.Workbook ):
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 3, max_row = 3, values_only = True ) ) )
    labelindex = { label: index for index, label in enumerate( labels ) }
    labelindex[None] += 1 # (make final past-the-end column value)
    index = { index: label for label, index in labelindex.items() }

    rownr = 6
    for r in sheet.iter_rows( min_row = rownr, max_row = sheet.max_row, values_only = True ):
        if not r[0]:
            continue
        else:
            month = r[0]
        
        personnr = 0
        col = 1
        while( col < labelindex[None] ):
            if col in index:
                personnr = index[col]
                if not personnr:
                    break
            projectnr = r[col]
            if not projectnr:
                col += 3
                continue
            if r[col+1] not in ['P','B','F','O']:
                raise RuntimeError( f"problem in row {rownr}, column {col+1}: phase column needs to contain single character 'P', 'B', 'F', or 'O'" )
            phase = expenditure_phase( r[col+1] )
            amount = r[col+2]
            uniqblurb = '-'.join( [ str(projectnr), str(month), str(personnr) ] )
            if phase not in planning:
                planning[phase] = {}
            planning[phase][uniqblurb] = SalaryCost(
                phase     = phase,
                personnr  = personnr,
                date      = month,
                fraction  = 1,
                projectnr = projectnr,
            )
            planning[phase][uniqblurb].additem( nature   = 'Loonschatting',
                                                employee = personnr,
                                                date     = month,
                                                value    = amount,
                                               )
            col += 3
        rownr += 1
    
def parse_projectbudgets( projects: dict,
                          sheet: openpyxl.workbook.workbook.Workbook ):
    # Get the label header line in the 'rapport' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 1, max_row = 1, values_only = True ) ) )
    labelindex = {label: index for index, label in enumerate( labels ) }
    
    for r in sheet.iter_rows( min_row = 2, max_row = sheet.max_row, values_only = True ):
        project = r[labelindex['Project']]
        if not project in projects:
            raise RuntimeError( f"project unknown. Insert the following line on the Projects tab in Projects.xlsx:\n" +
                                f"{project};{r[labelindex['Proj.naam']]};N;;{r[labelindex['Promotor']]};{r[labelindex['Copromotor']]};{r[labelindex['Afd']]};{r[labelindex['Eenheid']]};{r[labelindex['Finbron']]};<fill-out>;{r[labelindex['Einddatum project']]}" )

        bi = BudgetItem( r[labelindex['Totaal Budget']],
                         r[labelindex['Vastgelegd']],
                         r[labelindex['Uitgegeven']] )
        projects[project].budget[ExpenditureType[r[labelindex['Uitgaven Categorie']]]] = bi;


def parse_budgetboeking( projects: dict,
                         sheet: openpyxl.workbook.workbook.Workbook ):
    # Get the label header line in the 'rapport' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 4, max_row = 4, values_only = True ) ) )
    # The line below does not work as a moron has given two columns the samen name
    # labelindex = { label: index for index, label in enumerate( labels ) }
    labelindex = {}
    for index, label in enumerate( labels ):
        while label in labelindex:
            label += '*'
        labelindex[label] = index
    
    for r in sheet.iter_rows( min_row = 5, max_row = sheet.max_row, values_only = True ):
        project   = r[labelindex['Activiteit/Project']]
        name      = r[labelindex['Activiteit/Project Omschrijving']]
        afdeling  = r[labelindex['Afdeling']]
        eenheid   = r[labelindex['Eenheid']]
        begindate = r[labelindex['Datum Transactie']]
        budget    = r[labelindex['Budget']]
        simplebudget : str
        if re.search( r"^FFI", project ):
            simplebudget = 'N'
        elif re.search( r"^FFP", project ):
            simplebudget = 'Y'
        else:
            raise RuntimeError( "internal error: could not classify project having a simple cost-budget or a complex budget (I only need Budgetboekingen-* for FFI or FFP projects)" )

        if not project in projects:
            raise RuntimeRror( "insert the following line on the Projects tab in Projects.xlsx:\n" +
                               "{project};{name};{simplebudget};;;;{afdeling};{eenheid};-;{begindate};-" )

        budget_type = r[labelindex['Rekening']]
        budget_item = BudgetItem( budget, 0, 0 )
        if budget_type == 'XKOSTEN':
            # onverdeeld budget
            projects[project].budget = budget_item
        elif budget_type == 'XWEDDEN':
            # wedden budget
            projects[project].budget[ExpenditureType.WEDDEN] = budget_item
        elif budget_type == 'XWERKING':
            projects[project].budget[ExpenditureType.WERKING] = budget_item
        elif budget_type == 'XINKOMST':
            # niet interessant = betaling door bedrijf
            continue
        else:
            raise RuntimeError( "internal error: could not determine budget line" )
        

def parse_salaries( data   : dict[ ExpenditureType, dict[ ExpenditurePhase, dict] ],
                    sheet: openpyxl.workbook.workbook.Workbook,
                    persons: dict[ str, Person ] ):
    salaries = data[ExpenditureType.WEDDEN]

    # Get the label header line in the 'rapport' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 1, max_row = 1, values_only = True ) ) )
    labelindex = {label: index for index, label in enumerate( labels ) }

    for r in sheet.iter_rows( min_row = 2, max_row = sheet.max_row, values_only = True ):
        budgetniveau = r[labelindex['B-NIV']]
        if budgetniveau == 'BW':
            continue
        month = re.sub( r"^(\d{4})(\d{2})$", "\\1-\\2", str( r[labelindex['Wedde-maand']] ) )
        projectnr = re.sub( r"^OZ", '', r[labelindex['Activiteit']] )
        personnr  = str( r[labelindex['Empl.nr']] or '00000' )
        person    = str( r[labelindex['Naam']] or 'N.N.' )
        uniqblurb = '-'.join( [ projectnr, month, personnr ] )
        if personnr not in persons:
            raise RuntimeError( f"person {person} with number {personnr} not found" )
        amount_bk = float(r[labelindex['Actuals']])
        amount_fx = float(r[labelindex['Vastleggingen']])
        phase: ExpenditurePhase
        amount: float
        if amount_fx != 0:
            phase = ExpenditurePhase.FIXED
            amount = amount_fx
        elif amount_bk != 0:
            phase = ExpenditurePhase.BOOKED
            amount = amount_bk
        else:
            if amount_fx == 0 and amount_bk == 0:
                raise RuntimeError( "internal issue: salary that is neither a booking, nor a fixing" )
        if uniqblurb not in salaries[phase]:
            salaries[phase][uniqblurb] = SalaryCost(
                phase     = phase,
                personnr  = personnr,
                date      = month,
                fraction  = 1,
                projectnr = projectnr,
            )
        salaries[phase][uniqblurb].additem( nature   = r[labelindex['Rekening Omschr']],
                                            employee = personnr,
                                            date     = month,
                                            value    = amount,
                                           )
            
def parse_fl_transactions( data   : dict[ ExpenditureType, dict[ ExpenditurePhase, dict] ],
                           sheet: openpyxl.workbook.workbook.Workbook,
                           persons: dict[ str, Person ],
                           projects ):
    
    # Get the label header line in the 'rapport' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 4, max_row = 4, values_only = True ) ) )

    # The line below does not work as a moron has given two columns the samen name
    # labelindex = { label: index for index, label in enumerate( labels ) }
    labelindex = {}
    for index, label in enumerate( labels ):
        while label in labelindex:
            label += '*'
        labelindex[label] = index

    # Now parse all rows
    for r in sheet.iter_rows( min_row = 5, max_row = sheet.max_row, values_only = True ):
        projectnr = r[labelindex['Activiteit']]
        
        rubriek  = r[labelindex['Rubriek']]
        sign: int
        if rubriek == 'Beginsaldo':
            if not projectnr in projects:
                raise RuntimeError( f"insert the following line on the Projects tab in Projects.xlsx and complete it:\n" + 
                                    "{projectnr};;Y;;;;42;{r[labelindex['Eenheid']]};Intern;01-01-2014;31-12-2099")
            projects[projectnr].register_budget( ExpenditureType.WERKING, r[labelindex['Bedrag']] )
            continue
        elif rubriek == 'Inkomsten':
            sign = -1
        elif rubriek == 'Uitgaven':
            sign = +1
        else:
            raise RuntimeException( f"internal problem: strange value in 'Rubriek' {rubriek}" )
                
        phase = r[labelindex['Subrubriek']]
        if phase == 'Actuals':
            phase = ExpenditurePhase.BOOKED
        elif phase == 'Vastleggingen huidig boekjaar':
            phase = ExpenditurePhase.FIXED
        else:
            raise RuntimeException( f"internal problem: did not expect this value: {phase} in 'Rubriek'" )

        type = ExpenditureType.WERKING
        if re.search( r"^62", str( r[labelindex['Rekening']] ) ):
             type = ExpenditureType.WEDDEN

        day = parse_day( r[labelindex['Datum Transactie']] )
        month = parse_month( r[labelindex['Datum Transactie']] )
        nature   = r[labelindex['Omschrijving']]         
        if type == ExpenditureType.WEDDEN:
            pass # the expenditures are taken from the general salaries transactions list
            # batchnr = str( r[labelindex['DocumentNr. /Bestelaanvr.' ]] )
            # uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
            # if uniqblurb not in data[type][phase]:
            #     data[type][phase][uniqblurb] = SalaryCost(
            #         phase     = phase,
            #         personnr  = '0',
            #         date      = month,
            #         fraction  = 1,
            #         projectnr = projectnr,
            #     )
            # data[type][phase][uniqblurb].additem( nature   = r[labelindex['Omschrijving']],
            #                                       employee = 'NN',
            #                                       date     = day,
            #                                       value    = r[labelindex['Bedrag']] )
        else:
            batchnr: str
            klant: str
            factuurnr: str
            if sign == 1:
                batchnr = str( r[labelindex['VoucherNr. /OntvangstNr.' ]] )
                klant = str( r[labelindex['Klant/Leverancier']] )
                factuurnr = str( r[labelindex['FactuurNr. Klant/Lev']] )
            else:
                batchnr = str( r[labelindex['DocumentNr. /Bestelaanvr.']] )
                klant = '-'
                factuurnr = '-'
            uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
            if uniqblurb not in data[type][phase]:
                data[type][phase][uniqblurb] = OperationCost( phase     = phase,
                                                              projectnr = projectnr,
                                                              batchnr   = batchnr,
                                                              date      = day,
                                                              nature    = r[labelindex['Omschrijving']] )
            data[type][phase][uniqblurb].additem( nature      = r[labelindex['Omschrijving']],
                                                  opponent    = klant,
                                                  date        = day,
                                                  value       = sign * r[labelindex['Bedrag']],
                                                  description = factuurnr )

def parse_ff_transactions( data   : dict[ ExpenditureType, dict[ ExpenditurePhase, dict] ],
                           sheet: openpyxl.workbook.workbook.Workbook,
                           persons: dict[ str, Person ],
                           projects ):
    
    # Get the label header line in the 'rapport' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 4, max_row = 4, values_only = True ) ) )

    # The line below does not work as a moron has given two columns the samen name
    # labelindex = { label: index for index, label in enumerate( labels ) }
    labelindex = {}
    for index, label in enumerate( labels ):
        while label in labelindex:
            label += '*'
        labelindex[label] = index

    for r in sheet.iter_rows( min_row = 5, max_row = sheet.max_row, values_only = True ):

        if r[labelindex['Category name']] in ( 'LONEN', None ):
            continue
            # the salaries are taken from the general salaries transactions list
            
        type = ExpenditureType.WERKING

        projectnr = r[labelindex['Activiteit/Project']]

        amount_bk = float(r[labelindex['Actual']])
        amount_fx = float(r[labelindex['Vastlegging']])

        phase: ExpenditurePhase
        amount: float
        if amount_fx != 0:
            phase = ExpenditurePhase.FIXED
            amount = amount_fx
        elif amount_bk != 0:
            phase = ExpenditurePhase.BOOKED
            amount = amount_bk
        else:
            if amount_fx == 0 and amount_bk == 0:
                raise RuntimeException( "internal problem: salary that is neither a booking, nor a fixing" )
        

        day = parse_day( r[labelindex['Datum Transactie']] )
        month = parse_month( r[labelindex['Datum Transactie']] )
        nature   = r[labelindex['Omschrijving']]         

        batchnr = str( r[labelindex['VoucherNr. /OntvangstNr.' ]] )
        klant = str( r[labelindex['Klant/Leverancier']] )
        factuurnr = str( r[labelindex['FactuurNr.  Klant/Lev.']] )
        uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
        
        if uniqblurb not in data[type][phase]:
            data[type][phase][uniqblurb] = OperationCost( phase     = phase,
                                                          projectnr = projectnr,
                                                          batchnr   = batchnr,
                                                          date      = day,
                                                          nature    = r[labelindex['Rekening Omschrijving']] )
        data[type][phase][uniqblurb].additem( nature      = nature,
                                              opponent    = klant,
                                              date        = day,
                                              value       = amount,
                                              description = factuurnr )
            

def parse_booked_expenditures( data   : dict[ ExpenditureType, dict[ ExpenditurePhase, dict] ],
                               sheet: openpyxl.workbook.workbook.Workbook,
                               persons: dict[ str, Person ] ):
    # We will build a list of expenditures, and create indices into that list, for
    # -
    equipment    = data[ExpenditureType.UITRUSTING][ExpenditurePhase.BOOKED]
    operations   = data[ExpenditureType.WERKING][ExpenditurePhase.BOOKED]
    overhead     = data[ExpenditureType.OVERHEAD][ExpenditurePhase.BOOKED]

    # Get the label header line in the 'rapport' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 1, max_row = 1, values_only = True ) ) )
    labelindex = {label: index for index, label in enumerate( labels ) }

    # Now parse all rows
    for r in sheet.iter_rows( min_row = 2, max_row = sheet.max_row, values_only = True ):
        type = r[labelindex['Uitgaven categorie']]
        if type  == ExpenditureType.WEDDEN.name:
            continue
        elif type == ExpenditureType.WERKING.name:
            day = parse_day( r[labelindex['Expend. Item Date']] )
            projectnr = r[labelindex['Project nr']]
            batchnr = r[labelindex['Vouchernr / Batchnr' ]]
            nature   = r[labelindex['Uitgaven soort']]         
            uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
            if uniqblurb not in operations:
                operations[uniqblurb] = OperationCost( phase = ExpenditurePhase.BOOKED,
                                                       projectnr = projectnr,
                                                       batchnr   = batchnr,
                                                       date      = day,
                                                       nature    = nature )
            operations[uniqblurb].additem( nature      = r[labelindex['Uitgaven soort']],
                                           opponent    = r[labelindex['Werknemer / Leverancier']],
                                           date        = day,
                                           value       = r[labelindex['Uitgegeven']],
                                           description = r[labelindex['Omschrijving']] )
        elif type == ExpenditureType.UITRUSTING.name:
            day = parse_day( r[labelindex['Expend. Item Date']] )
            projectnr = r[labelindex['Project nr']]
            batchnr = r[labelindex['Vouchernr / Batchnr' ]]
            nature   = r[labelindex['Uitgaven soort']]         
            uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
            if uniqblurb not in equipment:
                equipment[uniqblurb] = EquipmentCost( phase = ExpenditurePhase.BOOKED,
                                                       projectnr = projectnr,
                                                       batchnr   = batchnr,
                                                       date      = day,
                                                       nature    = nature )
            equipment[uniqblurb].additem( nature      = r[labelindex['Uitgaven soort']],
                                           opponent    = r[labelindex['Werknemer / Leverancier']],
                                           date        = day,
                                           value       = r[labelindex['Uitgegeven']],
                                           description = r[labelindex['Omschrijving']] )
        elif type == ExpenditureType.OVERHEAD.name:
            day = parse_day( r[labelindex['Expend. Item Date']] )
            projectnr = r[labelindex['Project nr']]
            batchnr = r[labelindex['Vouchernr / Batchnr' ]]
            nature   = r[labelindex['Uitgaven soort']]         
            uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
            if uniqblurb not in overhead:
                overhead[uniqblurb] = OverheadCost( phase = ExpenditurePhase.BOOKED,
                                                    projectnr = projectnr,
                                                    batchnr   = batchnr,
                                                    date      = day,
                                                   )
            overhead[uniqblurb].additem( nature      = r[labelindex['Uitgaven soort']],
                                         opponent    = str( r[labelindex['Werknemer / Leverancier']] ),
                                         date        = day,
                                         value       = r[labelindex['Uitgegeven']],
                                        )
        else:
            print( "strange!", type )

def parse_fixed_expenditures( data   : dict[ ExpenditureType, dict[ ExpenditurePhase, dict] ],
                              sheet: openpyxl.workbook.workbook.Workbook,
                              persons: dict[ str, Person ] ):
    # We will build a list of expenditures, and create indices into that list, for
    # -
    equipment    = data[ExpenditureType.UITRUSTING][ExpenditurePhase.FIXED]
    operations   = data[ExpenditureType.WERKING][ExpenditurePhase.FIXED]
    overhead     = data[ExpenditureType.OVERHEAD][ExpenditurePhase.FIXED]

    # Get the label header line in the 'rapport' worksheet
    labels = list( map( lambda x: x[0],
                        sheet.iter_cols( min_row = 1, max_row = 1, values_only = True ) ) )
    labelindex = {label: index for index, label in enumerate( labels ) }

    # Now parse all rows
    for r in sheet.iter_rows( min_row = 2, max_row = sheet.max_row, values_only = True ):
        type = r[labelindex['Uitgaven categorie']]
        if type  == ExpenditureType.WEDDEN.name:
            continue
        elif type == ExpenditureType.WERKING.name:
            day = parse_day( r[labelindex['Expend. Item Date']] )
            projectnr = r[labelindex['Project nr']]
            batchnr = str( r[labelindex['Vouchernr' ]] )
            nature = r[labelindex['Uitgaven soort']]
            uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
            if uniqblurb not in operations:
                operations[uniqblurb] = OperationCost( phase     = ExpenditurePhase.FIXED,
                                                       projectnr = projectnr,
                                                       batchnr   = batchnr,
                                                       date      = day,
                                                       nature    = nature,
                                                      )
            operations[uniqblurb].additem( nature      = r[labelindex['Uitgaven soort']],
                                           opponent    = r[labelindex['Leverancier']],
                                           date        = day,
                                           value       = r[labelindex['Vastgelegd']],
                                           description = r[labelindex['Omschrijving']],
                                          )
        elif type == ExpenditureType.UITRUSTING.name:
            day = parse_day( r[labelindex['Expend. Item Date']] )
            projectnr = r[labelindex['Project nr']]
            batchnr = r[labelindex['Vouchernr' ]]
            nature = r[labelindex['Uitgaven soort']]
            uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
            if uniqblurb not in equipment:
                equipment[uniqblurb] = EquipmentCost( phase     = ExpenditurePhase.FIXED,
                                                       projectnr = projectnr,
                                                       batchnr   = batchnr,
                                                       date      = day,
                                                       nature    = nature,
                                                      )
            equipment[uniqblurb].additem( nature      = r[labelindex['Uitgaven soort']],
                                           opponent    = r[labelindex['Leverancier']],
                                           date        = day,
                                           value       = r[labelindex['Vastgelegd']],
                                           description = r[labelindex['Omschrijving']],
                                          )
        elif type == ExpenditureType.OVERHEAD.name:
            day = parse_day( r[labelindex['Expend. Item Date']] )
            projectnr = r[labelindex['Project nr']]
            batchnr = str( r[labelindex['Vouchernr' ]] )
            nature = r[labelindex['Uitgaven soort']]
            uniqblurb = '-'.join( [ projectnr, day, batchnr ] )
            if uniqblurb not in overhead:
                overhead[uniqblurb] = OverheadCost( phase     = ExpenditurePhase.FIXED,
                                                    projectnr = projectnr,
                                                    batchnr   = batchnr,
                                                    date      = day,
                                                   )
            overhead[uniqblurb].additem( nature      = r[labelindex['Uitgaven soort']],
                                         opponent    = str( r[labelindex['Leverancier']] ),
                                         date        = day,
                                         value       = r[labelindex['Vastgelegd']],
                                        )
        else:
            print( "strange!", type )

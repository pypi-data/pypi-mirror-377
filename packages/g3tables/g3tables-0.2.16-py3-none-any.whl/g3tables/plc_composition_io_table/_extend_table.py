import argparse
import copy
import json
import logging
import os
import re
import typing

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

from ._table import (
    PLCCompositionIOTable as Table,
    PLCCompositionIOTableColumn as TableColumn
)


logger = logging.getLogger('g3tables.plc_composition_io_table')


def get_input_args():
    parser = argparse.ArgumentParser(
        description=(
            'Prefill the IO table with SW Device placeholders and '
            'the corresponding function block SW connector names.'
            )
        )
    parser.add_argument(
        'table_path',
        type=str,
        help='Path to the HW IO Table'
        )
    parser.add_argument(
        '-re',
        '--add-re-patterns',
        required=False,
        type=str,
        help='Path to the additional regex patterns file'
        )
    parser.add_argument(
        '--log-level',
        type=str,
        required=False,
        default='INFO',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
        help='Set the logging level.'
        )
    return parser.parse_args()


class IOSignalPattern(typing.TypedDict):
    module: str
    connector: str


def get_io_signal_patterns() -> dict[str, IOSignalPattern]:
    file_dir = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(file_dir, 'io_signals.json')
    logger.debug('Loading default IO mapping file "%s"', file_path)
    with open(file_path, 'r') as file:
        return json.load(file)


def get_user_io_signal_patterns(path: str) -> dict[str, IOSignalPattern]:
    logger.debug('Loading user IO mapping file "%s"', path)
    with open(path, 'r') as file:
        patterns = json.load(file)
    logger.debug('Validating user IO mapping patterns data')
    if not isinstance(patterns, dict):
        raise ValueError('Invalid pattern data structure.')
    for signal, data in patterns.items():
        logger.debug('Validating pattern "%s"', signal)
        try:
            module = data['module']
            logger.debug(
                'Retrieved "module=%s" for pattern "%s"', module, signal
                )
            connector = data['connector']
            logger.debug(
                'Retrieved "connector=%s" for pattern "%s"', connector, signal
                )
        except KeyError as err:
            raise ValueError('Invalid pattern data structure.') from err
    return patterns


IO_SIGNAL_PATTERNS = get_io_signal_patterns()


def get_table_name_and_extension(path: str) -> tuple[str, str]:
    logger.debug('Extracting table name and extension from "%s"', path)
    name, ext = os.path.splitext(os.path.basename(path))
    logger.debug('Extracted "name=%s" and "extension=%s"', name, ext)
    return name, ext


def get_table(path: str) -> Workbook:
    logger.debug('Loading "%s" with openpyxl', path)
    return load_workbook(path, data_only=True)


def get_column_index_and_letter(letter_or_index: str | int) -> tuple[int, str]:
    if isinstance(letter_or_index, str):
        logger.debug('Retrieving index for column "%s"', letter_or_index)
        index = column_index_from_string(letter_or_index)
        letter = letter_or_index.upper()
    else:
        logger.debug('Retrieving letter for column "%s"', letter_or_index)
        index = letter_or_index
        letter = get_column_letter(letter_or_index)
    logger.debug('Retrieved "index=%s" and "letter=%s"', index, letter)
    return index, letter


def get_colunm_names(sheet: Worksheet) -> list[str]:
    logger.debug('Extracting table sheet "%s" columns', sheet.title)
    names: list[str] = []
    for i in range(1, sheet.max_column + 1):
        name = sheet.cell(row=Table.HEADER_ROW, column=i).value
        name = str(name) if name else ''
        names.append(name)
    logger.debug('Extracted columns: %s', names)
    return names


def insert_column(
    sheet: Worksheet, index: int | str, name: str, width: int = 20
) -> None:
    logger.debug('Inserting column "%s" at index "%s"', name, index)
    index_int, index_str = get_column_index_and_letter(index)
    index_str_old = get_column_letter(index_int + 1)
    width_old = sheet.column_dimensions[index_str].width
    sheet.insert_cols(index_int)
    logger.debug('Setting column "%s" width to "%s"', index_str, width)
    sheet.column_dimensions[index_str].width = width
    sheet.column_dimensions[index_str_old].width = width_old
    logger.debug('Setting column "%s" name to "%s"', index_str, name)
    header = sheet.cell(row=Table.HEADER_ROW, column=index_int)
    header.value = name


def unmerge_cells_in_new_column(sheet: Worksheet, col_idx: int):
    """
    Must be placed after the formatting functions. For some reason,
    crashes if used before the formatting of both columns (KeyError).
    """
    affected_ranges = []
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col <= col_idx <= max_col and max_row > Table.HEADER_ROW:
            affected_ranges.append(merged_range)
    for merged_range in affected_ranges:
        sheet.unmerge_cells(str(merged_range))


def format_column(
    sheet: Worksheet, column_index_new: int, column_index_reference: int
) -> None:
    logger.debug(
        'Copying column cells\' formatting from "%s" to "%s"',
        get_column_letter(column_index_reference),
        get_column_letter(column_index_new)
        )
    for row in range(1, sheet.max_row + 1):
        cell_ref = sheet.cell(row=row, column=column_index_reference)
        cell_new = sheet.cell(row=row, column=column_index_new)
        if cell_ref.has_style:
            cell_new.font = copy.copy(cell_ref.font)
            cell_new.border = copy.copy(cell_ref.border)
            cell_new.number_format = copy.copy(cell_ref.number_format)
            cell_new.protection = copy.copy(cell_ref.protection)
            cell_new.alignment = copy.copy(cell_ref.alignment)
            column = Table.get_column_index(TableColumn.TERMINAL)
            if not sheet.cell(row=row, column=column).value:
                cell_new.fill = copy.copy(cell_ref.fill)
                continue


def extend_sheet(
    sheet: Worksheet,
    user_patterns: typing.Optional[dict[str, IOSignalPattern]] = None
) -> None:

    not_matched_symbol = "?"

    def is_cell_empty(cell: Cell | typing.Any) -> bool:
        return not cell.value or cell.value == not_matched_symbol

    def match_pattern(
        sw_device_type: Cell | typing.Any,
        sw_device_name: Cell | typing.Any,
        sw_signal: Cell | typing.Any,
        sw_signal_alias: Cell | typing.Any,
        patterns: dict[str, IOSignalPattern]
    ) -> bool:
        for pattern, data in patterns.items():
            matched = bool(re.search(pattern, str(sw_signal_alias.value)))
            if not matched:
                continue
            if (is_cell_empty(sw_device_type) and is_cell_empty(sw_signal)):
                module_name_full = data['module']
                module_name_last = module_name_full.split('/')[-1]
                sw_device_type.value = module_name_full
                sw_device_name.value = f"<{module_name_last} name>"
                sw_signal.value = data['connector']
            break
        return matched

    for row in range(Table.HEADER_ROW + 1, sheet.max_row + 1):
        io_name_cell = sheet.cell(
            row=row, column=Table.get_column_index(TableColumn.SIGNAL_NAME)
            )
        sw_device_type_cell = sheet.cell(
            row=row, column=Table.get_column_index(TableColumn.SW_DEVICE_TYPE)
            )
        sw_device_name_cell = sheet.cell(
            row=row, column=Table.get_column_index(TableColumn.SW_DEVICE_NAME)
            )
        sw_signal_cell = sheet.cell(
            row=row, column=Table.get_column_index(TableColumn.SW_SIGNAL)
            )
        signal_cell = sheet.cell(
            row=row, column=Table.get_column_index(TableColumn.SIGNAL)
            )
        if (
            not io_name_cell.value or
            not signal_cell.value or
            'safe' in str(io_name_cell.value).lower() or
            'reserve' in str(signal_cell.value).lower()
        ):
            continue
        matched = False
        if user_patterns:
            matched = match_pattern(
                sw_device_type_cell,
                sw_device_name_cell,
                sw_signal_cell,
                signal_cell,
                user_patterns
                )
        if not matched:
            matched = match_pattern(
                sw_device_type_cell,
                sw_device_name_cell,
                sw_signal_cell,
                signal_cell,
                IO_SIGNAL_PATTERNS
                )
        if not matched:
            if (
                is_cell_empty(sw_device_type_cell) and
                is_cell_empty(sw_signal_cell)
            ):
                sw_device_type_cell.value = not_matched_symbol
                sw_device_name_cell.value = f'<{not_matched_symbol}>'
                sw_signal_cell.value = not_matched_symbol


def extend_workbook(
    table: Workbook,
    user_patterns: typing.Optional[dict[str, IOSignalPattern]] = None,
) -> None:
    for sheet in table.worksheets:
        if sheet.title in Table.EXCLUDE_SHEETS:
            logger.debug('Ignoring table sheet "%s"', sheet.title)
            continue
        logger.debug('Extending table sheet "%s"', sheet.title)
        column_names = get_colunm_names(sheet)
        logger.debug(
            'Checking if column "%s" is present in the table', 'SWDeviceType'
            )
        # temp insert: the 'Cabinet' column is inserted by
        # PLCCompositionIOTable during table data parsing at runtime
        insert_column(
            sheet,
            Table.get_column_index(TableColumn.CABINET),
            'Temp'
            )
        if 'SWDeviceType' not in column_names:
            insert_column(
                sheet,
                Table.get_column_index(TableColumn.SW_DEVICE_TYPE),
                'SWDeviceType'
                )
            format_column(
                sheet,
                Table.get_column_index(TableColumn.SW_DEVICE_TYPE),
                Table.get_column_index(TableColumn.SW_DEVICE_TYPE) - 1,
                )
        logger.debug(
            'Checking if column "%s" is present in the table', 'SWDeviceName'
            )
        if 'SWDeviceName' not in column_names:
            insert_column(
                sheet,
                Table.get_column_index(TableColumn.SW_DEVICE_NAME),
                'SWDeviceName'
                )
            format_column(
                sheet,
                Table.get_column_index(TableColumn.SW_DEVICE_NAME),
                Table.get_column_index(TableColumn.SW_DEVICE_NAME) - 1,
                )
        logger.debug(
            'Checking if column "%s" is present in the table', 'SWSignal'
            )
        if 'SWSignal' not in column_names:
            insert_column(
                sheet,
                Table.get_column_index(TableColumn.SW_SIGNAL),
                'SWSignal'
                )
            format_column(
                sheet,
                Table.get_column_index(TableColumn.SW_SIGNAL),
                Table.get_column_index(TableColumn.SW_SIGNAL) - 1,
                )
        signal_col_letter = get_column_letter(
            Table.get_column_index(TableColumn.SIGNAL) - 1
            )  # -1 because the temp column before this column is deleted later
        remark_col_letter = get_column_letter(
            Table.get_column_index(TableColumn.REMARK) - 1
            )  # -1 because the temp column before this column is deleted later
        sheet.column_dimensions[signal_col_letter].width = 20
        sheet.column_dimensions[remark_col_letter].width = 60
        # NOTE: cannot place the unmerging before the formatting
        unmerge_cells_in_new_column(
            sheet, Table.get_column_index(TableColumn.SW_DEVICE_NAME)
            )
        unmerge_cells_in_new_column(
            sheet, Table.get_column_index(TableColumn.SW_DEVICE_TYPE)
            )
        unmerge_cells_in_new_column(
            sheet, Table.get_column_index(TableColumn.SW_SIGNAL)
            )
        extend_sheet(sheet, user_patterns)
        # remove the temp insert
        logger.debug(
            'Deleting column "%s" at index "%s"',
            'Temp', Table.get_column_index(TableColumn.CABINET)
            )
        sheet.delete_cols(1)


def extend_table(
    file_path: str,
    user_patterns_file_path: str | None = None,
    output_dir_path: str | None = None,
    output_file_name: str | None = None
) -> None:
    # load source table
    logger.info('Loading PLC Composition and IO Table "%s"', file_path)
    table = get_table(file_path)
    # load user regex patterns
    user_patterns: dict[str, IOSignalPattern] = {}
    if user_patterns_file_path:
        logger.info(
            'Extracting user IO mapping patterns from "%s"',
            user_patterns_file_path
            )
        user_patterns = get_user_io_signal_patterns(user_patterns_file_path)
    # extend the table columns
    logger.info('Extending PLC Composition and IO Table "%s"', file_path)
    extend_workbook(table, user_patterns)
    # save the updated table
    output_dir_path = output_dir_path or os.getcwd()
    if not output_file_name:
        table_name, table_ext = get_table_name_and_extension(file_path)
        output_file_name = f'{table_name} Extended{table_ext}'
    output_file_path = os.path.join(output_dir_path, output_file_name)
    logger.info(
        'Saving the extended PLC Composition and IO Table to "%s"',
        output_file_path
        )
    table.save(output_file_path)


def main() -> None:
    args = get_input_args()
    logging.basicConfig(
        level=args.log_level,
        format='[%(name)s] %(levelname)s:%(message)s'
        )
    extend_table(args.table_path, args.add_re_patterns)

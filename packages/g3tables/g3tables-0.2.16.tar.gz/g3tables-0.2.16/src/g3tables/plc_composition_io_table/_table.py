import enum
import logging
import numpy as np
import pandas as pd
import re
import typing

from openpyxl.utils import get_column_letter

from ..utils.gdrive import get_table_data_from_gdrive


logger = logging.getLogger('g3tables.plc_composition_io_table')


class PLCCompositionIOTableColumn(enum.StrEnum):
    """
    Enumeration of column names in a PLC Composition IO table. This enum
    facilitates easier and more readable access to specific columns
    within the table data.

    Each enum member represents a column name. The docstring for each member
    includes a brief description and indicates its position in both
    the original and the expanded table. The position is denoted in the format
    (original_index_numeric/original_index_alphabet,
    expanded_index_numeric/expanded_index_alphabet).
    """
    CABINET = "Cabinet"
    """
    The name of the cabinet, typically the same as the sheet name.
    This column exists only in the expanded table.
    Position in tables: Original (-), Expanded (1/A).
    """
    COUNT = "Count"
    """
    The count of PLC units in the cabinet. Originally, this is the first
    unnamed column. Position in tables: Original (1/A), Expanded (2/B).
    """
    PLC_UNIT = "PLC Unit / from left to right /"
    """
    The name of the PLC hardware unit.
    Position in tables: Original (2/B), Expanded (3/C).
    """
    TERMINAL = "Terminal at the module"
    """
    The terminal location at the PLC module.
    Position in tables: Original (3/C), Expanded (4/D).
    """
    LED_INDICATOR = "LED indicator at the module"
    """
    The LED indicator located on the PLC module.
    Position in tables: Original (4/D), Expanded (5/E).
    """
    SLOT_ST = "Slot ST"
    """
    The Slot ST identifier.
    Position in tables: Original (5/E), Expanded (6/F).
    """
    SIGNAL_NAME = "Signal name at PLC module"
    """
    The name of the signal at the PLC hardware module, such as
    'DigitalInput01'. Position in tables: Original (6/F), Expanded (7/G).
    """
    POLARITY = "Polarity of OUT /Pulse x or +24V for IN"
    """
    Describes the polarity for output signals or the voltage for input signals.
    Position in tables: Original (7/G), Expanded (8/H).
    """
    SW_DEVICE_TYPE = "SWDeviceType"
    """
    The type of the software device associated with a PLC hardware module
    signal, such as 'Cabinet/Fuse'. This column exists only in the expanded
    table. Position in tables: Original (-), Expanded (9/I).
    """
    SW_DEVICE_NAME = "SWDeviceName"
    """
    The name of the software device associated with a PLC hardware module
    signal, such as 'FU_01'. This column exists only in the expanded table.
    Position in tables: Original (-), Expanded (10/J).
    """
    SW_SIGNAL = "SWSignal"
    """
    The software signal mapped to a PLC hardware module signal, as defined in
    G3Core function block, such as 'inFuseOk'. This column exists only in
    the expanded table. Position in tables: Original (-), Expanded (11/K).
    """
    SIGNAL = "Signal"
    """
    The software signal mapped to a PLC hardware module signal, as shown on
    the electrical connections schema, such as 'FU1_ST'.
    Position in tables: Original (8/H), Expanded (12/L).
    """
    REMARK = "Remark"
    """
    Any comments or notes related to the signal mapping.
    Position in tables: Original (9/I), Expanded (13/M).
    """
    PLC_UNIT_SUFFIX = "PLCUnitSuffix"
    """
    A unique letter identifier for a plc module. Derived from the column
    "Terminal at the module" (the latter cannot be used directly due to
    the cases when one module has two assigned letters, e.g. X20cSI9100).
    Position in tables: Original (-), Expanded (14/N).
    * TODO: move this column to "C".
    """


PLCCompositionIOTableColumnIndex: dict[str, int] = {
    memb: i + 1 for i, memb in enumerate(PLCCompositionIOTableColumn)
    }


class PLCCompositionIOTable:
    """
    `PLCCompositionIOTable` is a representation of the G3 Project PLC
    Composition and IO table. It allows access to the data of the hardware
    cabinet sheets in the form of a `pandas.DataFrame`.

    Note that the "razitko", "TSC-NxG_T11", "TSC-NxG_T22", "TSC-NxG_T22_20",
    "template units", "units", "objedn", "List1", and "power" sheets are not
    represented within the dictionary.
    """
    NAME_PATTERN = 'plc_composition'
    """
    The pattern that the name of the SW Definition table must match to be
    automatically recognized in a Google Drive folder.
    """
    HEADER_ROW = 11
    """
    The row index of the header row in the table.
    """
    EXCLUDE_SHEETS = [
        'razitko',
        'TSC-NxG_T11',
        'TSC-NxG_T22',
        'TSC-NxG_T22_20',
        'template units',
        'units',
        'objedn',
        'List1',
        'power'
        ]
    """
    Sheets that are not represented in the data dictionary of
    a `PLCCompositionIOTable` instance.
    """

    def __init__(self, data: pd.DataFrame) -> None:
        """
        Standard initializer of a `PLCCompositionIOTable` instance. Although it
        can be used directly, alternative class constructors `from_local` and
        `from_gdrive` are adviced instead.

        Args:
            data (dict[str, pd.DataFrame]): a dict, where keys are\
            sheets names and values are the `pandas.DataFrame` objects\
            representing the sheet data.
        """
        self.data = data

    @classmethod
    def is_name_valid(cls, name: str) -> bool:
        """
        Check if the provided name matches the PLC Composition and IO table
        name pattern.

        Args:
            name (str): The name to validate.

        Returns:
            bool: True if the name is valid, False otherwise.
        """
        name_formatted = name.strip().replace(' ', '_').casefold()
        return cls.NAME_PATTERN in name_formatted

    @classmethod
    def from_local(
        cls, path: str, sheet_name: str | typing.Iterable[str]
    ) -> typing.Self:
        """
        Create a `PLCCompositionIOTable` instance from a local Excel file.

        Args:
            local_path (str): The local file path to the Excel file containing\
                the PLC Composition and IO table data.

        Returns:
            Self: An instance of `PLCCompositionIOTable` initialized\
                with data from the specified local Excel file.
        """
        if isinstance(sheet_name, str):
            sheet_name = [sheet_name]
        else:
            sheet_name = [name for name in sheet_name]
        logger.info(
            'Reading PLC Composition IO table cabitet "%s" data from local '
            'path: "%s"', ", ".join(sheet_name), path
            )
        sheets: dict[str, pd.DataFrame] = pd.read_excel(
            path,
            sheet_name,  # type: ignore
            skiprows=[idx for idx in range(cls.HEADER_ROW - 1)],
            )
        # check if additional columns are inserted
        to_rename = {'Unnamed: 0': PLCCompositionIOTableColumn.COUNT}
        for cabinet_name, sheet in sheets.items():
            PLCCompositionIOTableFormatter.check_inserted_cols(
                cabinet_name, sheet
                )
            sheet.rename(columns=to_rename, inplace=True)
        # format and concat sheet dict to one dataframe
        formatter = PLCCompositionIOTableFormatter.format_and_concat_sheets
        sheets_concat = formatter(sheets)
        return cls(sheets_concat)

    @classmethod
    def from_gdrive(
        cls, gdrive_table_name: str, sheet_name: str | typing.Iterable[str]
    ) -> typing.Self:
        """
        Create a `PLCCompositionIOTable` instance from a Google Drive file.

        Args:
            gdrive_table_name (str): The name of the table in Google Drive\
                containing the PLC Composition and IO table data.

        Returns:
            Self: An instance of `PLCCompositionIOTable` initialized\
                with data from the specified Google Drive file.
        """
        if isinstance(sheet_name, str):
            sheet_name = [sheet_name]
        logger.info(
            'Reading PLC Composition IO table cabitet "%s" data from Google '
            'Drive: "%s"', ", ".join(sheet_name), gdrive_table_name
            )
        try:
            sheets: dict[str, pd.DataFrame] = get_table_data_from_gdrive(
                gdrive_table_name,
                header_row=cls.HEADER_ROW,
                exclude_sheets=cls.EXCLUDE_SHEETS
                )
        except ValueError as err:
            logger.error('%s', str(err))
            exit()
        # check if additional columns are inserted
        to_rename = {'': PLCCompositionIOTableColumn.COUNT}
        for cabinet_name, sheet in sheets.items():
            PLCCompositionIOTableFormatter.check_inserted_cols(
                cabinet_name, sheet
                )
            sheet.rename(columns=to_rename, inplace=True)
        # format and concat sheet dict to one dataframe
        formatter = PLCCompositionIOTableFormatter.format_and_concat_sheets
        sheets_concat = formatter(sheets, include_only=sheet_name)
        return cls(sheets_concat)

    @classmethod
    def load(
        cls, path: str, sheet_name: str | typing.Iterable[str]
    ) -> typing.Self:
        """
        Load a SW Definition table from a local file or Google Drive.

        Args:
            path (str): Path to the SW Definition table file. If the path ends\
                with '.xlsx', the table is loaded from a local file.\
                Otherwise, the table is loaded from Google Drive.

        Returns:
            typing.Self: An instance of `SWDefinitionTable` initialized with\
                data from the specified file.
        """
        if path.endswith('.xlsx'):
            return cls.from_local(path, sheet_name)
        return cls.from_gdrive(path, sheet_name)

    @staticmethod
    def get_column_index(column: str | PLCCompositionIOTableColumn) -> int:
        """
        Get the index of a column in the PLC Composition and IO table.

        Args:
            column (str | PLCCompositionIOTableColumn): The column name or\
                enumeration member.

        Returns:
            int: The index of the column in the table.
        """
        column = PLCCompositionIOTableColumn(column)
        return PLCCompositionIOTableColumnIndex[column]

    @staticmethod
    def get_column_letter(column: str | PLCCompositionIOTableColumn) -> str:
        """
        Get the column letter of a column in the PLC Composition and IO table.

        Args:
            column (str | PLCCompositionIOTableColumn): The column name or\
                enumeration member.

        Returns:
            str: The column letter of the column in the table.
        """
        column = PLCCompositionIOTableColumn(column)
        return get_column_letter(PLCCompositionIOTableColumnIndex[column])

    @staticmethod
    def _get_plc_unit_type(row) -> str | None:
        unit_type = row[PLCCompositionIOTableColumn.PLC_UNIT]
        if PLCCompositionIOTableFormatter.is_plc_unit_type(unit_type):
            return unit_type
        return None

    @staticmethod
    def _get_plc_unit_name(
        unit_type: str, cabinet: str, name_suffix: str
    ) -> str:
        return f'{unit_type}_{cabinet}_{name_suffix}'

    @staticmethod
    def _extract_prefix(s: str) -> str:
        match = re.match(r'^[A-Za-z]+', s)
        return match.group(0) if match else ''

    def _is_terminal_equivalent(self, terminal1: str, terminal2: str) -> bool:
        prefix1 = self._extract_prefix(terminal1)
        prefix2 = self._extract_prefix(terminal2)
        return prefix1 == prefix2

    def iter_plc_units(self) -> typing.Iterator[tuple[str, str, str]]:
        """
        Iterate over the PLC units in the table.

        Yields:
            typing.Iterator[tuple[str, str, str]]: An iterator over the PLC\
                units in the table. Each iteration yields a tuple containing\
                the PLC unit type, the cabinet name, and the terminal suffix.
        """
        suffix_prev = ''
        for _, row in self.data.iterrows():
            plc_unit_type = self._get_plc_unit_type(row)
            if not plc_unit_type:
                continue
            suffix = row[PLCCompositionIOTableColumn.PLC_UNIT_SUFFIX]
            cabinet = row[PLCCompositionIOTableColumn.CABINET]
            if suffix == suffix_prev:
                continue
            else:
                suffix_prev = suffix
            yield plc_unit_type, cabinet, suffix

    def get_sl81xx_name(self) -> str | None:
        """
        Get the name of the "X20cSL81xx" PLC unit.

        Returns:
            str | None: The name of the "X20cSL81xx" PLC unit if found,\
                otherwise None.
        """
        logger.info('Searching for the "X20cSL81xx" PLC unit')
        for unit_type, cabinet, terminal in self.iter_plc_units():
            if 'sl8101' in unit_type.lower():
                sl8101 = self._get_plc_unit_name(unit_type, cabinet, terminal)
                logger.debug('Found the "X20cSL8101" PLC unit "%s"', sl8101)
                return sl8101
            if 'sl8100' in unit_type.lower():
                sl8100 = self._get_plc_unit_name(unit_type, cabinet, terminal)
                logger.debug('Found the "X20cSL8100" PLC unit "%s"', sl8100)
                return sl8100
        logger.warning('Could not find the "X20cSL81xx" PLC unit')
        return None

    def get_iomapping(self) -> dict[str, dict[str, str]]:
        """
        Get the IO mapping data from the PLC Composition and IO table.

        Returns:
            dict[str, dict[str, str]]: A dictionary containing the IO mapping\
                data. The keys are the software device type and name, and\
                the values are dictionaries containing the software signal\
                mapped to the hardware signal.
        """

        def get_i_prefix(signal_name: str) -> str:
            logger.debug(
                'Matching hardware signal "%s" IO mapping prefix', signal_name
                )
            name = signal_name.lower()
            if "digitalinput" in name:
                return "%IX"
            if "digitaloutput" in name:
                return "%QX"
            if "analoginput" in name:
                return "%IW"
            if "analogoutput" in name:
                return "%QW"
            if "temperature" in name:
                return "%IW"
            logger.warning('Unexpected hardware signal: "%s"', signal_name)
            return "%UNDEFINED_PREFIX"

        logger.info('Collecting IO mapping data')
        mapping: dict[str, dict[str, str]] = {}
        for _, row in self.data.iterrows():
            cabinet = row[PLCCompositionIOTableColumn.CABINET]
            sw_device_type = row[PLCCompositionIOTableColumn.SW_DEVICE_TYPE]
            if 'Cabinet/' in str(sw_device_type):
                logger.debug(
                    'Evaluated device "%s" as a Cabinet device or a Cabinet '
                    'child device', sw_device_type
                    )
                child_device_type = "/".join(sw_device_type.split('/')[1:])
                sw_device_type = f'Cabinet/{cabinet}/{child_device_type}'
            sw_device_name = row[PLCCompositionIOTableColumn.SW_DEVICE_NAME]
            sw_signal = row[PLCCompositionIOTableColumn.SW_SIGNAL]
            if not sw_device_type or not sw_device_name or not sw_signal:
                continue
            logger.debug(
                'Formatting IO mapping data for device "%s"', sw_device_type
                )
            name_suffix = self._extract_prefix(
                row[PLCCompositionIOTableColumn.TERMINAL]
                )
            hw_unit_type = row[PLCCompositionIOTableColumn.PLC_UNIT]
            hw_module_name = self._get_plc_unit_name(
                hw_unit_type, cabinet, name_suffix
                )
            hw_signal_name = row[PLCCompositionIOTableColumn.SIGNAL_NAME]
            prefix = get_i_prefix(hw_signal_name)
            hw_signal = f'{prefix}."{hw_module_name}".{hw_signal_name}'
            mapping_key = f'{sw_device_type}/{sw_device_name}'
            logger.debug(
                'Found IO mapping data for device "%s": "%s" -> "%s"',
                mapping_key, sw_signal, hw_signal
                )
            device_mapping = mapping.setdefault(mapping_key, dict())
            device_mapping[sw_signal] = hw_signal
        return mapping


class PLCCompositionIOTableFormatter:

    @staticmethod
    def check_inserted_cols(sheet_name: str, sheet: pd.DataFrame):
        cols = [
            PLCCompositionIOTableColumn.SW_DEVICE_TYPE,
            PLCCompositionIOTableColumn.SW_DEVICE_NAME,
            PLCCompositionIOTableColumn.SW_SIGNAL
        ]
        for col in cols:
            logger.info(
                'Checking if column "%s" is present in sheet "%s"',
                col, sheet_name
                )
            if col.value not in sheet.columns:
                raise AttributeError(
                    f'Column "{col.value}" was not found in '
                    f'table sheet "{sheet_name}".'
                    )

    @staticmethod
    def is_plc_unit_type(value) -> bool:
        if not value:
            return False
        value = str(value).strip()
        res = ('X20' in value) and value.isalnum()
        return res

    @staticmethod
    def format_to_int_str(val: typing.Any) -> str | typing.Any:
        # check if the value is numeric and not NaN
        if pd.notna(val) and np.issubdtype(type(val), np.number):
            return str(int(val))
        return val  # return the value unchanged if not numeric or is NaN

    @classmethod
    def add_plc_unit_suffix_col(cls, sheet: pd.DataFrame) -> pd.DataFrame:
        col_name_unit = PLCCompositionIOTableColumn.PLC_UNIT
        col_name_suff = PLCCompositionIOTableColumn.PLC_UNIT_SUFFIX
        col_name_term = PLCCompositionIOTableColumn.TERMINAL
        # get mask of the col values where row does not contain plc unit name
        is_plc_unit_mask = sheet[col_name_unit].apply(cls.is_plc_unit_type)
        # format the suffix column
        sheet[col_name_suff] = None
        term_letters = sheet.loc[is_plc_unit_mask, col_name_term]
        sheet.loc[is_plc_unit_mask, col_name_suff] = term_letters
        sheet[col_name_suff] = sheet[col_name_suff].ffill()
        return sheet

    @classmethod
    def format_plc_unit_col(cls, sheet: pd.DataFrame) -> pd.DataFrame:
        col_name = PLCCompositionIOTableColumn.PLC_UNIT
        # get mask of the col values where row does not contain plc unit name
        is_plc_unit_mask = sheet[col_name].apply(cls.is_plc_unit_type)
        # set non-module plc unit cells to None
        sheet.loc[~is_plc_unit_mask, col_name] = None
        # forward fill the module names
        sheet[col_name] = sheet[col_name].ffill()
        return sheet

    @classmethod
    def format_sheet(cls, sheet: pd.DataFrame) -> pd.DataFrame:
        logger.info('Dropping all empty rows')
        sheet = sheet.replace('', np.nan)
        sheet = sheet.dropna(axis=0, how='all')
        logger.debug(
            'Updating "%s" column values',
            PLCCompositionIOTableColumn.PLC_UNIT
            )
        sheet = cls.add_plc_unit_suffix_col(sheet)
        sheet = cls.format_plc_unit_col(sheet)
        logger.debug('Converting all cell values string representation')
        sheet = sheet.applymap(cls.format_to_int_str)
        sheet = sheet.where(pd.notna(sheet), '')
        sheet = sheet.fillna('')
        sheet = sheet.applymap(lambda cell: str(cell).strip())
        return sheet

    @classmethod
    def format_and_concat_sheets(
        cls,
        sheets: dict[str, pd.DataFrame],
        ignore: typing.Iterable[str] | None = None,
        include_only: typing.Iterable[str] | None = None
    ) -> pd.DataFrame:
        if ignore is None:
            ignore = []
        if include_only is None:
            include_only = [name for name in sheets.keys()]
        sheets_concat = pd.DataFrame()
        logger.info('Formatting and concatenating cabinet sheets')
        for cabinet_name, sheet in sheets.items():
            if cabinet_name in ignore or cabinet_name not in include_only:
                logger.debug('Ignoring sheet "%s"', cabinet_name)
                continue
            logger.info('Formatting sheet "%s"', cabinet_name)
            sheet = cls.format_sheet(sheet)
            # insert 'Cabinet' column
            logger.info(
                'Insering column "%s" into sheet "%s"',
                PLCCompositionIOTableColumn.CABINET, cabinet_name
                )
            sheet.insert(1, 'Cabinet', cabinet_name)
            # concatenate sheets
            logger.info('Concatenting sheet "%s"', cabinet_name)
            to_concat = [sheets_concat, sheet]
            sheets_concat = pd.concat(to_concat, ignore_index=True)
        return sheets_concat

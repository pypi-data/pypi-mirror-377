from __future__ import annotations

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl import _ZipFileFileProtocol


def filter_warnings():
    import warnings
    warnings.filterwarnings("ignore", module="openpyxl.*")


def excel2json(
        io: _ZipFileFileProtocol,
        sheet_name: str | None = None,
        header: int = 1,
        warnings: bool = True
    ) -> list[dict]:
    from openpyxl import load_workbook
    from io import BytesIO
    if not warnings:
        filter_warnings()

    wb = load_workbook(BytesIO(io) if isinstance(io, bytes) else io)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=header, max_row=header))]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=header+1, values_only=True)]

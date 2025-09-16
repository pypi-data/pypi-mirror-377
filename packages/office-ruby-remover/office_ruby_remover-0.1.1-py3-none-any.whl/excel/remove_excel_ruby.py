"""
Excel files からルビ（ふりがな）を削除するモジュール
"""
from __future__ import annotations
import os
import logging
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# 定数定義
EXCEL_NAMESPACE = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NAMESPACES = {'x': EXCEL_NAMESPACE}
SUCCESS = 0
FAILURE = 1

# ロガーの設定
logger = logging.getLogger(__name__)

class ExcelProcessingError(Exception):
    """Excel処理中のエラーを示すカスタム例外"""
    pass

def validate_paths(input_path: str | Path, output_path: str | Path, overwrite: bool = False) -> None:
    """
    入力・出力パスのバリデーションを行う

    Args:
        input_path: 入力ファイルのパス
        output_path: 出力ファイルのパス
        overwrite: 上書きを許可するかどうか

    Raises:
        ExcelProcessingError: パスのバリデーションに失敗した場合
    """
    input_path = Path(input_path)
    output_path = Path(output_path)

    if not input_path.exists():
        raise ExcelProcessingError("入力ファイルが存在しません")
    
    if output_path.exists() and not overwrite:
        raise ExcelProcessingError("出力ファイルが既に存在し、上書きが許可されていません")

def process_ruby_elements(cell: Cell | MergedCell) -> None:
    """
    セル内のルビ要素を処理する

    Args:
        cell: 処理対象のセル

    Note:
        セルにアクセスできない場合やXML操作が不可能な場合は何もせずにスキップ
    """
    if not cell.value:  # 空のセルはスキップ
        return

    try:
        element = getattr(cell, '_element', None)  # type: ignore
        if element is None:
            return

        # ルビ要素を検索
        ruby_elements = element.findall('.//x:ruby', NAMESPACES)
        for ruby in ruby_elements:
            # ベーステキスト（ルビなしテキスト）を取得
            base_text = ruby.find('.//x:rubyBase//x:t', NAMESPACES)
            if base_text is not None:
                # ルビ要素をベーステキストで置換
                ruby.getparent().replace(ruby, base_text)
    except AttributeError:
        # XML操作をサポートしていないセル（結合セルなど）はスキップ
        logger.debug(f"セルの処理をスキップしました: {cell.coordinate}")

def process_worksheet(worksheet: Worksheet) -> None:
    """
    ワークシート内の全セルのルビを処理する

    Args:
        worksheet: 処理対象のワークシート
    """
    for row in worksheet.iter_rows():
        for cell in row:
            process_ruby_elements(cell)

def remove_excel_ruby(input_path: str | Path, output_path: str | Path, overwrite: bool = False) -> int:
    """
    Excelファイルからルビ（ふりがな）を削除する

    Args:
        input_path: 入力Excelファイルのパス
        output_path: 保存先Excelファイルのパス
        overwrite: 出力ファイルの上書きを許可するかどうか

    Returns:
        int: ステータスコード（0: 成功, 1: 失敗）

    Note:
        - 入力ファイルは変更されません
        - ルビが含まれていないファイルも正常に処理されます
    """
    try:
        # パスのバリデーション
        validate_paths(input_path, output_path, overwrite)

        # ワークブックを読み込み
        workbook = load_workbook(str(input_path))
        try:
            # 各ワークシートを処理
            for worksheet in workbook.worksheets:
                process_worksheet(worksheet)

            # 変更を保存
            workbook.save(str(output_path))
            logger.info(f"ルビを削除し、ファイルを保存しました: {output_path}")
            return SUCCESS

        finally:
            workbook.close()

    except ExcelProcessingError as e:
        logger.error(str(e))
        return FAILURE
    except Exception as e:
        logger.exception(f"予期しないエラーが発生しました: {e}")
        return FAILURE

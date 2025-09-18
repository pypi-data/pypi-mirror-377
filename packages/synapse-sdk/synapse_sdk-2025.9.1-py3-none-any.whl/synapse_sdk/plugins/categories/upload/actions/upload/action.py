import asyncio
import os
import shutil
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Awaitable, Dict, List, Optional, TypeVar

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from synapse_sdk.clients.exceptions import ClientError
from synapse_sdk.clients.utils import get_batched_list
from synapse_sdk.clients.validators.collections import FileSpecificationValidator
from synapse_sdk.plugins.categories.base import Action
from synapse_sdk.plugins.categories.decorators import register_action
from synapse_sdk.plugins.categories.upload.actions.upload.models import UploadParams
from synapse_sdk.plugins.enums import PluginCategory, RunMethod
from synapse_sdk.plugins.exceptions import ActionError
from synapse_sdk.utils.storage import get_pathlib

from .enums import LogCode, UploadStatus
from .exceptions import ExcelParsingError, ExcelSecurityError
from .run import UploadRun
from .utils import ExcelMetadataUtils, ExcelSecurityConfig

T = TypeVar('T')


@register_action
class UploadAction(Action):
    """Main upload action for processing and uploading files to storage.

    Handles file upload operations including validation, file discovery,
    Excel metadata processing, and data unit generation. Supports both
    synchronous and asynchronous upload processing with comprehensive
    progress tracking and error handling.

    Features:
    - Recursive directory scanning
    - Excel metadata validation and processing
    - Batch processing for large file sets
    - Type-based file organization
    - Security validation for Excel files
    - Progress tracking with detailed metrics
    - Comprehensive error logging

    Class Attributes:
        name (str): Action identifier ('upload')
        category (PluginCategory): UPLOAD category
        method (RunMethod): JOB execution method
        run_class (type): UploadRun for specialized logging
        params_model (type): UploadParams for parameter validation
        progress_categories (dict): Progress tracking configuration
        metrics_categories (dict): Metrics collection configuration

    Example:
        >>> action = UploadAction(
        ...     params={
        ...         'name': 'Data Upload',
        ...         'path': '/data/files',
        ...         'storage': 1,
        ...         'collection': 5
        ...     },
        ...     plugin_config=config
        ... )
        >>> result = action.run_action()
    """

    name = 'upload'
    category = PluginCategory.UPLOAD
    method = RunMethod.JOB
    run_class = UploadRun
    params_model = UploadParams
    progress_categories = {
        'analyze_collection': {
            'proportion': 2,
        },
        'upload_data_files': {
            'proportion': 38,
        },
        'generate_data_units': {
            'proportion': 60,
        },
    }
    metrics_categories = {
        'data_files': {
            'stand_by': 0,
            'failed': 0,
            'success': 0,
        },
        'data_units': {
            'stand_by': 0,
            'failed': 0,
            'success': 0,
        },
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.excel_config = ExcelSecurityConfig()
        self.excel_utils = ExcelMetadataUtils(self.excel_config)

    def get_uploader(self, path, file_specification, organized_files, params: Dict = {}):
        """Get uploader from entrypoint."""
        return self.entrypoint(
            self.run, path, file_specification, organized_files, extra_params=params.get('extra_params')
        )

    def _discover_files_recursive(self, dir_path: Path) -> List[Path]:
        return [file_path for file_path in dir_path.rglob('*') if file_path.is_file()]

    def _discover_files_non_recursive(self, dir_path: Path) -> List[Path]:
        return [file_path for file_path in dir_path.glob('*') if file_path.is_file()]

    def _validate_excel_security(self, excel_path: Path) -> None:
        file_size = excel_path.stat().st_size
        if file_size > self.excel_config.MAX_FILE_SIZE_BYTES:
            raise ExcelSecurityError(
                f'Excel file too large: {file_size} bytes (max: {self.excel_config.MAX_FILE_SIZE_BYTES})'
            )

        estimated_memory = file_size * 3
        if estimated_memory > self.excel_config.MAX_MEMORY_USAGE_BYTES:
            raise ExcelSecurityError(
                f'Excel file may consume too much memory: ~{estimated_memory} bytes '
                f'(max: {self.excel_config.MAX_MEMORY_USAGE_BYTES})'
            )

    def _prepare_excel_file(self, excel_path: Path) -> BytesIO:
        self._validate_excel_security(excel_path)
        excel_bytes = excel_path.read_bytes()
        return BytesIO(excel_bytes)

    def _process_excel_headers(self, headers: tuple) -> tuple:
        if len(headers) < 2:
            raise ExcelParsingError('Excel file must have at least 2 columns (file name and metadata)')
        self._validate_excel_content(headers, 0)
        return headers

    def _process_excel_data_row(self, row: tuple, headers: tuple) -> Optional[Dict[str, Any]]:
        if not row[0] or str(row[0]).strip() == '':
            return None

        file_name = str(row[0]).strip()
        if not self.excel_utils.is_valid_filename_length(file_name):
            self.run.log_message_with_code(LogCode.FILENAME_TOO_LONG, file_name[:50])
            return None

        file_metadata: Dict[str, Any] = {}
        for i, value in enumerate(row[1:], start=1):
            if value is not None and i < len(headers):
                header_value = headers[i]
                column_name = str(header_value).strip() if header_value is not None else f'column_{i}'

                column_name = self.excel_utils.validate_and_truncate_string(
                    column_name, self.excel_config.MAX_COLUMN_NAME_LENGTH
                )
                str_value = self.excel_utils.validate_and_truncate_string(
                    str(value), self.excel_config.MAX_METADATA_VALUE_LENGTH
                )
                file_metadata[column_name] = str_value

        return {file_name: file_metadata} if file_metadata else None

    def _process_excel_worksheet(self, worksheet) -> Dict[str, Dict[str, Any]]:
        if worksheet is None:
            raise ExcelParsingError('Excel file has no active worksheet')

        metadata_dict: Dict[str, Dict[str, Any]] = {}
        headers: Optional[tuple] = None
        data_row_count = 0
        validation_interval = getattr(self.excel_config, 'VALIDATION_CHECK_INTERVAL', 1000)

        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            if row_idx == 0:
                headers = self._process_excel_headers(row)
                continue

            if headers is None:
                raise ExcelParsingError('Excel file missing header row')

            data_row_count += 1

            if data_row_count % validation_interval == 0:
                self._validate_excel_content(headers, data_row_count)

            row_result = self._process_excel_data_row(row, headers)
            if row_result:
                metadata_dict.update(row_result)

        self._validate_excel_content(headers or (), data_row_count)

        return metadata_dict

    def _validate_excel_content(self, headers: tuple, row_count: int) -> None:
        if len(headers) > self.excel_config.MAX_COLUMNS:
            raise ExcelParsingError(f'Too many columns: {len(headers)} (max: {self.excel_config.MAX_COLUMNS})')

        if row_count > self.excel_config.MAX_ROWS:
            raise ExcelParsingError(f'Too many rows: {row_count} (max: {self.excel_config.MAX_ROWS})')

    def _find_excel_metadata_file(self, pathlib_cwd: Path) -> Optional[Path]:
        for extension in ['.xlsx', '.xls']:
            excel_path = pathlib_cwd / f'meta{extension}'
            if excel_path.exists() and excel_path.is_file():
                return excel_path
        return None

    def _read_excel_metadata(self, pathlib_cwd: Path) -> Dict[str, Dict[str, Any]]:
        excel_path = None

        excel_metadata_path = self.params.get('excel_metadata_path')
        if excel_metadata_path:
            excel_path = pathlib_cwd / excel_metadata_path
            if not excel_path.exists():
                self.run.log_message_with_code(LogCode.EXCEL_FILE_NOT_FOUND_PATH)
                return {}
        else:
            excel_path = self._find_excel_metadata_file(pathlib_cwd)
            if not excel_path:
                return {}

        try:
            self.run.log_message_with_code(LogCode.EXCEL_FILE_VALIDATION_STARTED)

            excel_stream = self._prepare_excel_file(excel_path)

            workbook = load_workbook(excel_stream, read_only=True, data_only=True)
            try:
                self.run.log_message_with_code(LogCode.EXCEL_WORKBOOK_LOADED)
                return self._process_excel_worksheet(workbook.active)
            finally:
                workbook.close()

        except ExcelSecurityError as e:
            self.run.log_message_with_code(LogCode.EXCEL_SECURITY_VALIDATION_FAILED, str(e))
            raise
        except ExcelParsingError as e:
            self.run.log_message_with_code(LogCode.EXCEL_PARSING_FAILED, str(e))
            raise
        except InvalidFileException as e:
            self.run.log_message_with_code(LogCode.EXCEL_INVALID_FILE_FORMAT, str(e))
            raise ExcelParsingError(f'Invalid Excel file format: {str(e)}')
        except MemoryError:
            self.run.log_message_with_code(LogCode.EXCEL_FILE_TOO_LARGE)
            raise ExcelSecurityError('Excel file exceeds memory limits')
        except (OSError, IOError) as e:
            self.run.log_message_with_code(LogCode.EXCEL_FILE_ACCESS_ERROR, str(e))
            raise ExcelParsingError(f'File access error: {str(e)}')
        except Exception as e:
            self.run.log_message_with_code(LogCode.EXCEL_UNEXPECTED_ERROR, str(e))
            raise ExcelParsingError(f'Unexpected error: {str(e)}')

    def start(self) -> Dict[str, Any]:
        result: Dict[str, Any] = {}

        storage_id = self.params.get('storage')
        if storage_id is None:
            raise ActionError('Storage parameter is required')
        storage = self.client.get_storage(storage_id)

        path = self.params.get('path')
        if path is None:
            raise ActionError('Path parameter is required')
        pathlib_cwd = get_pathlib(storage, path)

        excel_metadata: Dict[str, Dict[str, Any]] = {}
        try:
            excel_metadata = self._read_excel_metadata(pathlib_cwd)
            if excel_metadata:
                self.run.log_message_with_code(LogCode.EXCEL_METADATA_LOADED, len(excel_metadata))
        except ExcelSecurityError as e:
            self.run.log_message_with_code(LogCode.EXCEL_SECURITY_VIOLATION, str(e))
            return result
        except ExcelParsingError as e:
            if self.params.get('excel_metadata_path'):
                self.run.log_message_with_code(LogCode.EXCEL_PARSING_ERROR, str(e))
                return result
            else:
                self.run.log_message_with_code(LogCode.EXCEL_PARSING_ERROR, str(e))
                excel_metadata = {}

        file_specification_template = self._analyze_collection()
        organized_files = self._organize_files(pathlib_cwd, file_specification_template, excel_metadata)

        uploader = self.get_uploader(pathlib_cwd, file_specification_template, organized_files)

        organized_files = uploader.handle_upload_files()

        if not self._validate_organized_files(organized_files, file_specification_template):
            self.run.log_message_with_code(LogCode.VALIDATION_FAILED)
            raise ActionError('Upload is aborted due to validation errors.')

        if not organized_files:
            self.run.log_message_with_code(LogCode.NO_FILES_FOUND)
            raise ActionError('Upload is aborted due to missing files.')

        if self.params.get('use_async_upload', True):
            uploaded_files = self.run_async(self._upload_files_async(organized_files, 10))
        else:
            uploaded_files = self._upload_files(organized_files)
        result['uploaded_files_count'] = len(uploaded_files)

        if not uploaded_files:
            self.run.log_message_with_code(LogCode.NO_FILES_UPLOADED)
            raise ActionError('Upload is aborted due to no uploaded files.')
        generated_data_units = self._generate_data_units(
            uploaded_files, self.params.get('creating_data_unit_batch_size', 1)
        )
        result['generated_data_units_count'] = len(generated_data_units)

        if not generated_data_units:
            self.run.log_message_with_code(LogCode.NO_DATA_UNITS_GENERATED)
            raise ActionError('Upload is aborted due to no generated data units.')

        self._cleanup_temp_directory()

        self.run.log_message_with_code(LogCode.IMPORT_COMPLETED)
        return result

    def _analyze_collection(self) -> Dict[str, Any]:
        self.run.set_progress(0, 2, category='analyze_collection')

        collection_id = self.params.get('data_collection')
        if collection_id is None:
            raise ActionError('Data collection parameter is required')
        self.run.set_progress(1, 2, category='analyze_collection')

        collection = self.run.client.get_data_collection(collection_id)
        self.run.set_progress(2, 2, category='analyze_collection')

        return collection['file_specifications']

    def _upload_files(self, organized_files: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        organized_files_count = len(organized_files)
        self.run.set_progress(0, organized_files_count, category='upload_data_files')
        self.run.log_message_with_code(LogCode.UPLOADING_DATA_FILES)

        client = self.run.client
        collection_id = self.params.get('data_collection')
        if collection_id is None:
            raise ActionError('Data collection parameter is required')
        upload_result = []
        current_progress = 0
        success_count = 0
        failed_count = 0

        self._update_metrics(organized_files_count, success_count, failed_count, 'data_files')

        for organized_file in organized_files:
            try:
                use_chunked_upload = self._requires_chunked_upload(organized_file)
                uploaded_data_file = client.upload_data_file(organized_file, collection_id, use_chunked_upload)
                self.run.log_data_file(organized_file, UploadStatus.SUCCESS)
                success_count += 1
                upload_result.append(uploaded_data_file)
            except Exception as e:
                self.run.log_data_file(organized_file, UploadStatus.FAILED)
                self.run.log_message_with_code(LogCode.FILE_UPLOAD_FAILED, str(e))
                failed_count += 1

            current_progress += 1
            self._update_metrics(organized_files_count, success_count, failed_count, 'data_files')
            self.run.set_progress(current_progress, organized_files_count, category='upload_data_files')

        self.run.set_progress(organized_files_count, organized_files_count, category='upload_data_files')

        return upload_result

    def run_async(self, coro: Awaitable[T]) -> T:
        import concurrent.futures

        def _run_in_thread():
            return asyncio.run(coro)

        try:
            asyncio.get_running_loop()
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(_run_in_thread)
                return future.result()
        except RuntimeError:
            return asyncio.run(coro)

    async def _upload_files_async(
        self, organized_files: List[Dict[str, Any]], max_concurrent: int = 10
    ) -> List[Dict[str, Any]]:
        organized_files_count = len(organized_files)
        self.run.set_progress(0, organized_files_count, category='upload_data_files')
        self.run.log_message_with_code(LogCode.UPLOADING_DATA_FILES)

        client = self.run.client
        collection_id = self.params.get('data_collection')
        if collection_id is None:
            raise ActionError('Data collection parameter is required')
        upload_result = []
        success_count = 0
        failed_count = 0

        self._update_metrics(organized_files_count, success_count, failed_count, 'data_files')

        semaphore = asyncio.Semaphore(max_concurrent)

        async def upload_single_file(organized_file):
            async with semaphore:
                loop = asyncio.get_event_loop()
                try:
                    use_chunked_upload = self._requires_chunked_upload(organized_file)
                    uploaded_data_file = await loop.run_in_executor(
                        None, lambda: client.upload_data_file(organized_file, collection_id, use_chunked_upload)
                    )
                    self.run.log_data_file(organized_file, UploadStatus.SUCCESS)
                    return {'status': 'success', 'result': uploaded_data_file}
                except ClientError as e:
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code(LogCode.FILE_UPLOAD_FAILED, f'Client error: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'client_error', 'retryable': True}
                except (OSError, IOError) as e:
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code(LogCode.FILE_UPLOAD_FAILED, f'File system error: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'file_error', 'retryable': False}
                except MemoryError as e:
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code(
                        LogCode.FILE_UPLOAD_FAILED, f'Memory error (file too large): {str(e)}'
                    )
                    return {'status': 'failed', 'error': str(e), 'error_type': 'memory_error', 'retryable': False}
                except asyncio.TimeoutError as e:
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code(LogCode.FILE_UPLOAD_FAILED, f'Upload timeout: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'timeout_error', 'retryable': True}
                except ValueError as e:
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code(LogCode.FILE_UPLOAD_FAILED, f'Data validation error: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'validation_error', 'retryable': False}
                except Exception as e:
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code(LogCode.FILE_UPLOAD_FAILED, f'Unexpected error: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'unknown_error', 'retryable': False}

        tasks = [upload_single_file(organized_file) for organized_file in organized_files]

        current_progress = 0
        for completed_task in asyncio.as_completed(tasks):
            result = await completed_task
            current_progress += 1

            if result['status'] == 'success':
                success_count += 1
                upload_result.append(result['result'])
            else:
                failed_count += 1

            self._update_metrics(organized_files_count, success_count, failed_count, 'data_files')
            self.run.set_progress(current_progress, organized_files_count, category='upload_data_files')

        self.run.set_progress(organized_files_count, organized_files_count, category='upload_data_files')

        return upload_result

    def _generate_data_units(self, uploaded_files: List[Dict[str, Any]], batch_size: int) -> List[Dict[str, Any]]:
        upload_result_count = len(uploaded_files)
        self.run.set_progress(0, upload_result_count, category='generate_data_units')
        self.run.log_message_with_code(LogCode.GENERATING_DATA_UNITS)

        client = self.run.client
        generated_data_units = []
        current_progress = 0
        success_count = 0
        failed_count = 0

        batches = get_batched_list(uploaded_files, batch_size)
        batches_count = len(batches)

        self._update_metrics(upload_result_count, success_count, failed_count, 'data_units')

        for batch in batches:
            try:
                created_data_units = client.create_data_units(batch)
                success_count += len(created_data_units)
                generated_data_units.append(created_data_units)
                for created_data_unit in created_data_units:
                    self.run.log_data_unit(
                        created_data_unit['id'], UploadStatus.SUCCESS, data_unit_meta=created_data_unit.get('meta')
                    )
            except Exception as e:
                failed_count += len(batch)
                self.run.log_message_with_code(LogCode.DATA_UNIT_BATCH_FAILED, str(e))
                for _ in batch:
                    self.run.log_data_unit(None, UploadStatus.FAILED, data_unit_meta=None)

            current_progress += 1
            self._update_metrics(upload_result_count, success_count, failed_count, 'data_units')
            self.run.set_progress(current_progress, batches_count, category='generate_data_units')

        self.run.set_progress(upload_result_count, upload_result_count, category='generate_data_units')

        return sum(generated_data_units, [])

    def _validate_organized_files(
        self, organized_files: List[Dict[str, Any]], file_specification_template: Dict[str, Any]
    ) -> bool:
        validator = FileSpecificationValidator(file_specification_template, organized_files)
        return validator.validate()

    def _organize_files(
        self,
        directory: Path,
        file_specification: List[Dict[str, Any]],
        excel_metadata: Optional[Dict[str, Dict[str, Any]]] = None,
    ) -> List[Dict[str, Any]]:
        organized_files: List[Dict[str, Any]] = []

        type_dirs: Dict[str, Path] = {}

        for spec in file_specification:
            spec_name = spec['name']
            spec_dir = directory / spec_name
            if spec_dir.exists() and spec_dir.is_dir():
                type_dirs[spec_name] = spec_dir

        if type_dirs:
            self.run.log_message_with_code(LogCode.TYPE_DIRECTORIES_FOUND, list(type_dirs.keys()))

        if not type_dirs:
            self.run.log_message_with_code(LogCode.NO_TYPE_DIRECTORIES)
            return organized_files

        self.run.log_message_with_code(LogCode.TYPE_STRUCTURE_DETECTED)
        self.run.log_message_with_code(LogCode.FILE_ORGANIZATION_STARTED)

        dataset_files = {}
        required_specs = [spec['name'] for spec in file_specification if spec.get('is_required', False)]

        is_recursive = self.params.get('is_recursive', True)

        for spec_name, dir_path in type_dirs.items():
            if is_recursive:
                files_list = self._discover_files_recursive(dir_path)
            else:
                files_list = self._discover_files_non_recursive(dir_path)

            for file_path in files_list:
                file_name = file_path.stem

                if file_name not in dataset_files:
                    dataset_files[file_name] = {}

                if spec_name not in dataset_files[file_name]:
                    dataset_files[file_name][spec_name] = file_path
                else:
                    existing_file = dataset_files[file_name][spec_name]
                    if file_path.stat().st_mtime > existing_file.stat().st_mtime:
                        dataset_files[file_name][spec_name] = file_path

        if not dataset_files:
            self.run.log_message_with_code(LogCode.NO_FILES_FOUND_WARNING)
            return organized_files

        self.run.log_message_with_code(LogCode.FILES_DISCOVERED, len(dataset_files))

        for file_name, files_dict in sorted(dataset_files.items()):
            if all(req in files_dict for req in required_specs):
                file_extensions = {}
                for file_path in files_dict.values():
                    ext = file_path.suffix.lower()
                    if ext:
                        file_extensions[ext] = file_extensions.get(ext, 0) + 1

                origin_file_extension = max(file_extensions.items(), key=lambda x: x[1])[0] if file_extensions else ''

                meta_data: Dict[str, Any] = {
                    'origin_file_stem': file_name,
                    'origin_file_extension': origin_file_extension,
                    'created_at': datetime.now().isoformat(),
                }

                if excel_metadata and file_name in excel_metadata:
                    meta_data.update(excel_metadata[file_name])

                organized_files.append({'files': files_dict, 'meta': meta_data})
            else:
                missing = [req for req in required_specs if req not in files_dict]
                self.run.log_message_with_code(LogCode.MISSING_REQUIRED_FILES, file_name, ', '.join(missing))

        return organized_files

    def _get_file_size_mb(self, file_path: Path) -> float:
        return file_path.stat().st_size / (1024 * 1024)

    def _requires_chunked_upload(self, organized_file: Dict[str, Any]) -> bool:
        max_file_size_mb = self.params.get('max_file_size_mb', 50)
        for file_path in organized_file.get('files', {}).values():
            if isinstance(file_path, Path) and self._get_file_size_mb(file_path) > max_file_size_mb:
                return True
        return False

    def _cleanup_temp_directory(self, temp_path: Optional[Path] = None) -> None:
        if temp_path is None:
            try:
                temp_path = Path(os.getcwd()) / 'temp'
            except (FileNotFoundError, OSError):
                return

        if not temp_path.exists():
            return

        shutil.rmtree(temp_path, ignore_errors=True)
        self.run.log_message(f'Cleaned up temporary directory: {temp_path}')

    def _update_metrics(self, total_count: int, success_count: int, failed_count: int, category: str):
        if not self.run:
            raise ValueError('Run instance not properly initialized')

        assert isinstance(self.run, UploadRun)

        metrics = self.run.MetricsRecord(
            stand_by=total_count - success_count - failed_count, failed=failed_count, success=success_count
        )
        self.run.log_metrics(metrics, category)
